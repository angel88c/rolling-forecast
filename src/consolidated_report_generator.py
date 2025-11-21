"""
Generador de reportes consolidados usando el template.
Llena el template con datos de Forecast, Costo de Venta y KPIs PM-008.
"""

import pandas as pd
import openpyxl
from openpyxl.styles import numbers
from datetime import datetime
from typing import Dict, Optional
import logging
from io import BytesIO
import calendar

logger = logging.getLogger(__name__)


class ConsolidatedReportGenerator:
    """Genera reportes consolidados llenando el template con datos procesados."""
    
    def __init__(self, template_path: str = 'data/Template.xlsx'):
        """
        Inicializa el generador de reportes.
        
        Args:
            template_path: Ruta al archivo template
        """
        self.template_path = template_path
        # TESTING agrupa ICT+FCT, OTROS agrupa TRN+SWD
        # REP se incluye como BU individual (no forma parte de OTROS)
        self.bus = ['TODAS', 'TESTING', 'ICT', 'FCT', 'IAT', 'REP', 'OTROS']
        self.locations = ['SAPI', 'LLC']
    
    def generate_report(
        self,
        forecast_results: Optional[Dict] = None,
        kpi_results: Optional[Dict] = None
    ) -> BytesIO:
        """
        Genera el reporte consolidado completo.
        
        Args:
            forecast_results: Resultados del procesamiento de forecast (ambos >=60% y <60%)
            kpi_results: Resultados del procesamiento de KPIs PM-008
            
        Returns:
            BytesIO con el Excel generado
        """
        logger.info("Iniciando generación de reporte consolidado")
        
        # Cargar template
        wb = openpyxl.load_workbook(self.template_path)
        ws = wb['Hoja2']
        
        # Extraer meses únicos de los datos fuente
        months = self._extract_months_from_data(forecast_results, kpi_results)
        
        if not months:
            logger.warning("No se encontraron meses en los datos")
            months = []
        
        logger.info(f"Meses extraídos de datos fuente: {len(months)}")
        
        # Escribir headers de meses en la fila 1
        self._write_month_headers(ws, months)
        
        # Llenar datos por cada BU
        current_row = 2  # Empezar desde la fila 2
        
        for bu in self.bus:
            logger.info(f"Procesando BU: {bu}")
            
            # 1. Facturación por Backlog PM (KPIs)
            self._fill_kpi_billing(ws, current_row, 'SAPI', bu, months, kpi_results)
            current_row += 1
            self._fill_kpi_billing(ws, current_row, 'LLC', bu, months, kpi_results)
            current_row += 1
            
            # 2. Costo de venta por Backlog PM (KPIs)
            self._fill_kpi_cost(ws, current_row, 'SAPI', bu, months, kpi_results)
            current_row += 1
            self._fill_kpi_cost(ws, current_row, 'LLC', bu, months, kpi_results)
            current_row += 1
            
            # 3. Facturación por Forecast >= 60%
            self._fill_forecast_billing(ws, current_row, 'SAPI', bu, months, forecast_results, low_prob=False)
            current_row += 1
            self._fill_forecast_billing(ws, current_row, 'LLC', bu, months, forecast_results, low_prob=False)
            current_row += 1
            
            # 4. Facturación por Forecast < 60%
            self._fill_forecast_billing(ws, current_row, 'SAPI', bu, months, forecast_results, low_prob=True)
            current_row += 1
            self._fill_forecast_billing(ws, current_row, 'LLC', bu, months, forecast_results, low_prob=True)
            current_row += 1
            
            # 5. Costo de venta por Forecast >= 60%
            self._fill_forecast_cost(ws, current_row, 'SAPI', bu, months, forecast_results, low_prob=False)
            current_row += 1
            self._fill_forecast_cost(ws, current_row, 'LLC', bu, months, forecast_results, low_prob=False)
            current_row += 1
            
            # 6. Costo de venta por Forecast < 60%
            self._fill_forecast_cost(ws, current_row, 'SAPI', bu, months, forecast_results, low_prob=True)
            current_row += 1
            self._fill_forecast_cost(ws, current_row, 'LLC', bu, months, forecast_results, low_prob=True)
            current_row += 1
        
        # Aplicar formato de moneda a las celdas de datos (columnas dinámicas según cantidad de meses)
        end_col = 3 + len(months)  # Columna D (4) es la primera, entonces 3 + cantidad de meses
        self._apply_currency_format(ws, start_row=2, end_row=current_row-1, start_col=4, end_col=end_col)
        
        # Guardar en BytesIO
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        
        logger.info("Reporte consolidado generado exitosamente")
        return buffer
    
    def _extract_months_from_data(self, forecast_results: Optional[Dict], kpi_results: Optional[Dict]) -> list:
        """
        Extrae todos los meses únicos de los datos fuente.
        
        Args:
            forecast_results: Resultados de forecast
            kpi_results: Resultados de KPIs
            
        Returns:
            Lista de strings con nombres de meses ordenados cronológicamente
        """
        month_set = set()
        
        # Extraer de forecast_table
        if forecast_results and 'forecast_table' in forecast_results:
            df = forecast_results['forecast_table'].get('data', [])
            if df:
                df = pd.DataFrame(df)
                # Buscar columnas que parecen meses (formato "Month YYYY")
                for col in df.columns:
                    if self._is_month_column(col):
                        month_set.add(str(col))
        
        # Extraer de forecast_table_low_prob
        if forecast_results and 'forecast_table_low_prob' in forecast_results:
            df = forecast_results['forecast_table_low_prob'].get('data', [])
            if df:
                df = pd.DataFrame(df)
                for col in df.columns:
                    if self._is_month_column(col):
                        month_set.add(str(col))
        
        # Extraer de cost_of_sale_table
        if forecast_results and 'cost_of_sale_table' in forecast_results:
            df = forecast_results['cost_of_sale_table'].get('data', [])
            if df:
                df = pd.DataFrame(df)
                for col in df.columns:
                    if self._is_month_column(col):
                        month_set.add(str(col))
        
        # Extraer de cost_of_sale_table_low_prob
        if forecast_results and 'cost_of_sale_table_low_prob' in forecast_results:
            df = forecast_results['cost_of_sale_table_low_prob'].get('data', [])
            if df:
                df = pd.DataFrame(df)
                for col in df.columns:
                    if self._is_month_column(col):
                        month_set.add(str(col))
        
        # Extraer de billing_table (KPIs)
        if kpi_results and 'billing_table' in kpi_results:
            df = kpi_results['billing_table'].get('data', [])
            if df:
                df = pd.DataFrame(df)
                for col in df.columns:
                    if self._is_month_column(col):
                        month_set.add(str(col))
        
        # Extraer de cost_table (KPIs)
        if kpi_results and 'cost_table' in kpi_results:
            df = kpi_results['cost_table'].get('data', [])
            if df:
                df = pd.DataFrame(df)
                for col in df.columns:
                    if self._is_month_column(col):
                        month_set.add(str(col))
        
        # Convertir a lista y ordenar cronológicamente
        months_list = sorted(list(month_set), key=self._parse_month_string)
        
        return months_list
    
    def _is_month_column(self, col_name: str) -> bool:
        """
        Determina si una columna es un mes basado en su nombre.
        
        Args:
            col_name: Nombre de la columna
            
        Returns:
            True si parece ser un mes
        """
        if not isinstance(col_name, str):
            return False
        
        # Buscar patrones como "November 2025", "Nov 2025", "January 2026"
        month_names = ['January', 'February', 'March', 'April', 'May', 'June',
                      'July', 'August', 'September', 'October', 'November', 'December',
                      'Jan', 'Feb', 'Mar', 'Apr', 'May', 'Jun', 
                      'Jul', 'Aug', 'Sep', 'Oct', 'Nov', 'Dec']
        
        for month in month_names:
            if month in col_name and any(char.isdigit() for char in col_name):
                return True
        
        return False
    
    def _parse_month_string(self, month_str: str) -> datetime:
        """
        Convierte string de mes a datetime para ordenamiento.
        
        Args:
            month_str: String con formato "Month YYYY" o "Mon YYYY"
            
        Returns:
            datetime object
        """
        formats = ['%B %Y', '%b %Y', '%B %y', '%b %y']
        
        for fmt in formats:
            try:
                return datetime.strptime(month_str, fmt)
            except:
                continue
        
        # Si no se puede parsear, retornar fecha muy lejana
        return datetime(2099, 12, 31)
    
    def _write_month_headers(self, ws, months: list):
        """
        Escribe los headers de meses en la fila 1 del Excel.
        
        Args:
            ws: Worksheet de openpyxl
            months: Lista de strings con nombres de meses
        """
        # Escribir columnas fijas primero (A, B, C)
        ws.cell(row=1, column=1, value='FACTURACION')
        ws.cell(row=1, column=2, value='Location')
        ws.cell(row=1, column=3, value='BU')
        
        # Escribir meses desde columna D (4) en adelante
        for idx, month_str in enumerate(months, start=4):
            # Convertir a datetime para formato consistente
            month_date = self._parse_month_string(month_str)
            ws.cell(row=1, column=idx, value=month_date)
    
    def _fill_kpi_billing(self, ws, row: int, location: str, bu: str, months: list, kpi_results: Optional[Dict]):
        """Llena datos de facturación por KPIs PM-008."""
        if not kpi_results or 'data' not in kpi_results:
            logger.debug(f"No hay datos de KPIs disponibles para billing (Location: {location}, BU: {bu})")
            return
        
        df_data = kpi_results.get('data', [])
        if not df_data:
            logger.debug(f"Lista de datos KPIs vacía (Location: {location}, BU: {bu})")
            return
        
        df = pd.DataFrame(df_data)
        logger.debug(f"KPI Billing - Location: {location}, BU: {bu}, Registros totales: {len(df)}")
        
        # Filtrar por Location y BU
        if bu == 'TODAS':
            filtered = df[df['Location'] == location]
        elif bu == 'TESTING':
            # TESTING incluye ICT, FCT
            testing_bus = ['ICT', 'FCT']
            filtered = df[(df['Location'] == location) & (df['BU'].isin(testing_bus))]
        elif bu == 'OTROS':
            # OTROS incluye TRN, SWD (REP es BU individual)
            otros_bus = ['TRN', 'SWD']
            filtered = df[(df['Location'] == location) & (df['BU'].isin(otros_bus))]
        else:
            filtered = df[(df['Location'] == location) & (df['BU'] == bu)]
        
        logger.debug(f"KPI Billing filtrados - Location: {location}, BU: {bu}, Registros: {len(filtered)}")
        
        # Sumar por mes
        self._fill_monthly_data(ws, row, filtered, months)
    
    def _generate_kpi_cost_table(self, df_billing: pd.DataFrame, months: list) -> pd.DataFrame:
        """
        Genera tabla de costo de venta a partir de la tabla de billing de KPIs.
        El costo se asigna al último mes con facturación.
        
        Args:
            df_billing: DataFrame con datos de billing
            months: Lista de meses
            
        Returns:
            DataFrame con tabla de costo
        """
        # Identificar columnas de meses en el DataFrame original
        month_cols = [col for col in df_billing.columns 
                     if col not in ['Proyecto', 'BU', 'Location', 'Status', 'Customer', 'Total PO', '% Facturación', 'Costo de Venta']]
        
        # Crear tabla base con información del proyecto
        df_cost = df_billing[['Proyecto', 'BU', 'Location', 'Status', 'Customer', 'Costo de Venta']].copy()
        
        # Inicializar columnas de meses con 0
        for month in month_cols:
            df_cost[month] = 0
        
        # Para cada proyecto, encontrar el último mes con facturación y asignar el costo ahí
        for idx in df_billing.index:
            last_month_with_billing = None
            
            # Buscar último mes con valor > 0
            for month in month_cols:
                if pd.notna(df_billing.loc[idx, month]) and df_billing.loc[idx, month] > 0:
                    last_month_with_billing = month
            
            # Si encontramos un mes con facturación, asignar el costo total ahí
            if last_month_with_billing and pd.notna(df_billing.loc[idx, 'Costo de Venta']):
                df_cost.loc[idx, last_month_with_billing] = df_billing.loc[idx, 'Costo de Venta']
        
        return df_cost
    
    def _fill_kpi_cost(self, ws, row: int, location: str, bu: str, months: list, kpi_results: Optional[Dict]):
        """Llena datos de costo de venta por KPIs PM-008."""
        if not kpi_results or 'data' not in kpi_results:
            logger.debug(f"No hay datos de KPIs disponibles para costo (Location: {location}, BU: {bu})")
            return
        
        df_data = kpi_results.get('data', [])
        if not df_data:
            logger.debug(f"Lista de datos KPIs vacía para costo (Location: {location}, BU: {bu})")
            return
        
        df_billing = pd.DataFrame(df_data)
        logger.debug(f"KPI Cost - Location: {location}, BU: {bu}, Registros billing: {len(df_billing)}")
        
        # Generar tabla de costo (igual que en app.py)
        df_cost = self._generate_kpi_cost_table(df_billing, months)
        logger.debug(f"Tabla de costo generada con {len(df_cost)} registros")
        
        # Filtrar por Location y BU
        if bu == 'TODAS':
            filtered = df_cost[df_cost['Location'] == location]
        elif bu == 'TESTING':
            # TESTING incluye ICT, FCT
            testing_bus = ['ICT', 'FCT']
            filtered = df_cost[(df_cost['Location'] == location) & (df_cost['BU'].isin(testing_bus))]
        elif bu == 'OTROS':
            # OTROS incluye TRN, SWD (REP es BU individual)
            otros_bus = ['TRN', 'SWD']
            filtered = df_cost[(df_cost['Location'] == location) & (df_cost['BU'].isin(otros_bus))]
        else:
            filtered = df_cost[(df_cost['Location'] == location) & (df_cost['BU'] == bu)]
        
        logger.debug(f"KPI Cost filtrados - Location: {location}, BU: {bu}, Registros: {len(filtered)}")
        
        # Sumar por mes
        self._fill_monthly_data(ws, row, filtered, months)
    
    def _fill_forecast_billing(self, ws, row: int, location: str, bu: str, months: list, 
                               forecast_results: Optional[Dict], low_prob: bool = False):
        """Llena datos de facturación por Forecast."""
        if not forecast_results:
            return
        
        # Seleccionar la tabla correcta
        if low_prob:
            table_key = 'forecast_table_low_prob'
        else:
            table_key = 'forecast_table'
        
        if table_key not in forecast_results:
            return
        
        df = forecast_results[table_key].get('data', [])
        if not df:
            return
        
        df = pd.DataFrame(df)
        
        # Filtrar por Empresa (Location) y BU
        if bu == 'TODAS':
            filtered = df[df['Empresa'] == location]
        elif bu == 'TESTING':
            # TESTING incluye ICT, FCT
            testing_bus = ['ICT', 'FCT']
            filtered = df[(df['Empresa'] == location) & (df['BU'].isin(testing_bus))]
        elif bu == 'OTROS':
            # OTROS incluye TRN, SWD (REP es BU individual)
            otros_bus = ['TRN', 'SWD']
            filtered = df[(df['Empresa'] == location) & (df['BU'].isin(otros_bus))]
        else:
            filtered = df[(df['Empresa'] == location) & (df['BU'] == bu)]
        
        # Sumar por mes
        self._fill_monthly_data(ws, row, filtered, months)
    
    def _fill_forecast_cost(self, ws, row: int, location: str, bu: str, months: list,
                           forecast_results: Optional[Dict], low_prob: bool = False):
        """Llena datos de costo de venta por Forecast."""
        if not forecast_results:
            return
        
        # Seleccionar la tabla correcta
        if low_prob:
            table_key = 'cost_of_sale_table_low_prob'
        else:
            table_key = 'cost_of_sale_table'
        
        if table_key not in forecast_results:
            return
        
        df = forecast_results[table_key].get('data', [])
        if not df:
            return
        
        df = pd.DataFrame(df)
        
        # Filtrar por Empresa (Location) y BU
        if bu == 'TODAS':
            filtered = df[df['Empresa'] == location]
        elif bu == 'TESTING':
            # TESTING incluye ICT, FCT
            testing_bus = ['ICT', 'FCT']
            filtered = df[(df['Empresa'] == location) & (df['BU'].isin(testing_bus))]
        elif bu == 'OTROS':
            # OTROS incluye TRN, SWD (REP es BU individual)
            otros_bus = ['TRN', 'SWD']
            filtered = df[(df['Empresa'] == location) & (df['BU'].isin(otros_bus))]
        else:
            filtered = df[(df['Empresa'] == location) & (df['BU'] == bu)]
        
        # Sumar por mes
        self._fill_monthly_data(ws, row, filtered, months)
    
    def _fill_monthly_data(self, ws, row: int, filtered_df: pd.DataFrame, months: list):
        """
        Llena los datos mensuales en la fila especificada.
        
        Args:
            ws: Worksheet de openpyxl
            row: Fila donde llenar datos
            filtered_df: DataFrame filtrado con los datos
            months: Lista de strings con nombres de meses exactos de las columnas
        """
        if filtered_df.empty:
            return
        
        # Columna inicial (D = 4)
        col_idx = 4
        
        for month_str in months:
            # Buscar columna exacta en el dataframe
            total = 0
            if month_str in filtered_df.columns:
                # Sumar valores de esa columna (excluyendo filas de totales)
                values = pd.to_numeric(filtered_df[month_str], errors='coerce').fillna(0)
                total = values.sum()
            
            # Escribir el valor en la celda
            ws.cell(row=row, column=col_idx, value=float(total) if total != 0 else None)
            col_idx += 1
    
    def _apply_currency_format(self, ws, start_row: int, end_row: int, start_col: int, end_col: int):
        """
        Aplica formato de moneda a un rango de celdas.
        
        Args:
            ws: Worksheet
            start_row: Fila inicial
            end_row: Fila final
            start_col: Columna inicial
            end_col: Columna final
        """
        currency_format = '$#,##0.00'
        
        for row_idx in range(start_row, end_row + 1):
            for col_idx in range(start_col, end_col + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                if cell.value is not None:
                    cell.number_format = currency_format
