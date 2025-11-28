"""
Clase base para managers de forecast.
Contiene lógica común de procesamiento y renderizado.
"""

import streamlit as st
import pandas as pd
from typing import List, Dict, Optional
from datetime import datetime
from io import BytesIO
import logging

from src.data_processor import DataProcessor
from src.validators import DataValidator
from src.forecast_calculator import ForecastCalculator
from src.ui_components import (
    AGGridConfigurator,
    GridResponseHandler,
    GRID_CONFIGS,
    create_section_header,
    render_filters_row,
    render_totals_panel,
    render_export_buttons
)
from src.formatters import format_currency as fmt_currency
from config.settings import INFO_MESSAGES, BUSINESS_RULES
from st_aggrid import AgGrid

logger = logging.getLogger(__name__)


class BaseForecastManager:
    """Clase base para gestionar diferentes tipos de forecast."""
    
    def __init__(self, sheet_name: str = 0):
        """
        Inicializa el manager con los componentes necesarios.
        
        Args:
            sheet_name: Nombre o índice de la hoja de Excel a procesar (default: 0)
        """
        self.processor = DataProcessor()
        self.validator = DataValidator()
        self.calculator = ForecastCalculator()
        self.sheet_name = sheet_name
    
    def merge_results_with_existing(self, new_results: Dict) -> Dict:
        """
        Hace merge de nuevos resultados con los existentes en session_state,
        preservando datos de otras pestañas.
        
        Args:
            new_results: Nuevos resultados a agregar/actualizar
            
        Returns:
            Dict con resultados combinados
        """
        # Si no hay resultados previos, retornar los nuevos
        if not hasattr(st.session_state, 'forecast_results'):
            return new_results
        
        # Obtener resultados existentes
        existing = st.session_state.forecast_results
        
        # Crear copia para no modificar el original
        merged = existing.copy()
        
        # Actualizar con nuevos valores (solo las claves que vienen en new_results)
        for key, value in new_results.items():
            merged[key] = value
        
        return merged
    
    def process_file(self, uploaded_file) -> Dict:
        """
        Procesa un archivo Excel de forecast.

        Args:
            uploaded_file: Archivo subido
            
        Returns:
            Dict con resultados del procesamiento
        """
        try:
            with st.spinner("Procesando forecast..."):
                # Paso 1: Validar archivo
                file_validation = self.validator.validate_file(uploaded_file)
                if not file_validation.is_valid:
                    st.error("❌ " + "; ".join(file_validation.errors))
                    return None
                
                # Paso 2: Leer archivo
                df, parsing_report = self.processor.read_excel_file(uploaded_file, sheet_name=self.sheet_name)
            
                # Verificar parsing exitoso
                if not parsing_report.get('parsing_success', False):
                    missing_cols = parsing_report.get('validation_result', {}).get('missing_columns', [])
                    if missing_cols:
                        st.error(f"❌ No se pudieron encontrar las siguientes columnas requeridas: {', '.join(missing_cols)}")
                        st.info("💡 Verifica que el archivo tenga las columnas necesarias")
                        return None
                
                # Paso 3: Procesar datos
                df_clean = self.processor.clean_and_prepare_data(df)

                # Paso 4: Validar datos procesados
                data_validation = self.validator.validate_dataframe(df_clean)
                
                # Mostrar advertencias
                if data_validation.warnings:
                    for warning in data_validation.warnings[:5]:
                        st.warning("⚠️ " + warning)
                
                # Paso 5: Convertir a objetos Opportunity
                opportunities_all = self.processor.convert_to_opportunities(df_clean)
                
                # Actualizar reglas de negocio
                self._update_business_rules()
                
                # Paso 6: Filtrar oportunidades según el tipo de manager
                opportunities = self.filter_opportunities(opportunities_all)
                
                if not opportunities:
                    st.warning(self.get_no_data_message())
                    return None
                
                # Mostrar info de filtrado
                self.show_filter_info(len(opportunities), len(opportunities_all))
                
                # Paso 7: Calcular forecast con tipo de facturación
                billing_type = getattr(st.session_state, 'billing_type', 'Contable')
                billing_events = self.calculator.calculate_forecast(opportunities, billing_type=billing_type)
                
                # Paso 8: Generar resumen y tablas
                summary = self.calculator.generate_forecast_summary(billing_events)
                forecast_table = self.calculator.create_forecast_table(billing_events)
                cost_of_sale_table = self.calculator.create_cost_of_sale_table(billing_events)
                
                # Paso 9: Preparar resultados
                results = self.prepare_results(
                    billing_events=billing_events,
                    summary=summary,
                    forecast_table=forecast_table,
                    cost_of_sale_table=cost_of_sale_table,
                    processing_summary=self.processor.get_processing_summary(df, df_clean, parsing_report),
                    validation_result=data_validation,
                    parsing_report=parsing_report,
                    opportunities_all=opportunities_all
                )
                
                st.success(self.get_success_message(len(opportunities)))
                
                return results
                
        except Exception as e:
            logger.error(f"Error en procesamiento: {str(e)}")
            st.error(f"❌ Error: {str(e)}")
            return None
    
    def filter_opportunities(self, opportunities: List) -> List:
        """
        Filtra oportunidades según el tipo de manager.
        Debe ser implementado por las clases hijas.
        
        Args:
            opportunities: Lista de todas las oportunidades
            
        Returns:
            Lista filtrada de oportunidades
        """
        raise NotImplementedError("Debe ser implementado por la clase hija")
    
    def prepare_results(self, **kwargs) -> Dict:
        """
        Prepara el diccionario de resultados.
        Debe ser implementado por las clases hijas.
        
        Returns:
            Dict con resultados formateados
        """
        raise NotImplementedError("Debe ser implementado por la clase hija")
    
    def get_no_data_message(self) -> str:
        """Retorna mensaje cuando no hay datos."""
        return "⚠️ No se encontraron oportunidades válidas"
    
    def get_success_message(self, count: int) -> str:
        """Retorna mensaje de éxito."""
        return f"✅ Forecast procesado: {count} oportunidades"
    
    def show_filter_info(self, filtered_count: int, total_count: int):
        """Muestra información sobre el filtrado aplicado."""
        pass  # Opcional, puede ser sobrescrito
    
    def _update_business_rules(self):
        """Actualiza reglas de negocio con valores editables del session state."""
        if hasattr(st.session_state, 'penalty_default'):
            BUSINESS_RULES.FINANCIAL_PENALTY_FACTOR_DEFAULT = st.session_state.penalty_default
            BUSINESS_RULES.FINANCIAL_PENALTY_FACTOR_60_PERCENT = st.session_state.penalty_60
            BUSINESS_RULES.INICIO_PERCENTAGE = st.session_state.inicio_pct
            BUSINESS_RULES.DR_PERCENTAGE = st.session_state.dr_pct
            BUSINESS_RULES.FAT_PERCENTAGE = st.session_state.fat_pct
            BUSINESS_RULES.SAT_PERCENTAGE = st.session_state.sat_pct
    
    def render_forecast_table(self, forecast_table: Dict, summary, title: str, key_prefix: str):
        """
        Renderiza la tabla de forecast con filtros y métricas.
        
        Args:
            forecast_table: Dict con datos de la tabla
            summary: Resumen del forecast
            title: Título de la sección
            key_prefix: Prefijo para keys de componentes
        """
        if len(forecast_table['data']) == 0:
            st.warning("No hay datos para mostrar")
            return
        
        df = pd.DataFrame(forecast_table['data'])
        
        # Filtros
        filter_configs = [
            {'column': 'Empresa', 'label': '🏢 Empresa', 'key': f'{key_prefix}_empresa'},
            {'column': 'BU', 'label': '📋 BU', 'key': f'{key_prefix}_bu'}
        ]
        
        col_filters, col_controls = st.columns([3, 1])
        
        with col_filters:
            selected_filters = render_filters_row(df, filter_configs)
        
        with col_controls:
            show_grouping = st.checkbox("Agrupar por BU", value=True, key=f"{key_prefix}_group")
        
        # Aplicar filtros
        df_filtered = df.copy()
        for column, value in selected_filters.items():
            if value != 'Todas':
                df_filtered = df_filtered[df_filtered[column] == value]
        
        # Mostrar panel de totales
        render_totals_panel(df_filtered, f"TOTALES {title.upper()}")
        
        # Configurar AG-Grid
        gb = AGGridConfigurator.configure_forecast_table(df_filtered)
        
        if selected_filters['BU'] != 'Todas' or not show_grouping:
            gb.configure_column("BU", rowGroup=False, hide=False)
            gb.configure_grid_options(groupDefaultExpanded=0)
        
        # Renderizar AG-Grid
        grid_config = GRID_CONFIGS['forecast_main'].copy()
        grid_config['height'] = AGGridConfigurator.get_grid_height(len(df_filtered), 600)
        
        st.markdown(f"#### 📊 Tabla de {title}")
        
        AgGrid(
            df_filtered,
            gridOptions=gb.build(),
            height=grid_config['height'],
            theme=grid_config['theme'],
            allow_unsafe_jscode=grid_config['allow_unsafe_jscode'],
            update_mode=grid_config['update_mode'],
            fit_columns_on_grid_load=grid_config['fit_columns_on_grid_load'],
            enable_enterprise_modules=grid_config['enable_enterprise_modules']
        )
        
        # Exportación
        st.markdown("#### 📥 Exportar Datos")
        render_export_buttons(df_filtered, key_prefix, f'{key_prefix}_export')
    
    def render_cost_of_sale_table(self, cost_table: Dict, summary, title: str, key_prefix: str):
        """
        Renderiza la tabla de costo de venta.
        
        Args:
            cost_table: Dict con datos de la tabla
            summary: Resumen del forecast
            title: Título de la sección
            key_prefix: Prefijo para keys de componentes
        """
        if len(cost_table['data']) == 0:
            st.warning("No hay datos de costo de venta disponibles")
            return
        
        df = pd.DataFrame(cost_table['data'])
        
        # Métricas principales
        col1, col2, col3 = st.columns(3)
        
        df_for_metrics = df[df['Proyecto'] != 'TOTAL COSTO'].copy() if 'TOTAL COSTO' in df['Proyecto'].values else df.copy()
        
        with col1:
            st.metric(f"📋 Proyectos {title}", len(df_for_metrics))
        
        with col2:
            if 'Costo de Venta' in df_for_metrics.columns:
                total_cost = pd.to_numeric(df_for_metrics['Costo de Venta'], errors='coerce').sum()
                st.metric(f"💸 Total Costo {title}", f"${total_cost:,.0f}")
        
        with col3:
            if 'Gross Margin' in df_for_metrics.columns:
                total_margin = pd.to_numeric(df_for_metrics['Gross Margin'], errors='coerce').sum()
                st.metric(f"💰 Total Gross Margin {title}", f"${total_margin:,.0f}")
        
        st.markdown("---")
        
        # Filtros
        filter_configs = [
            {'column': 'Empresa', 'label': '🏢 Empresa', 'key': f'{key_prefix}_cost_empresa'},
            {'column': 'BU', 'label': '📋 BU', 'key': f'{key_prefix}_cost_bu'}
        ]
        
        col_filters, col_controls = st.columns([3, 1])
        
        with col_filters:
            selected_filters = render_filters_row(df, filter_configs)
        
        with col_controls:
            show_grouping = st.checkbox("Agrupar por BU", value=True, key=f"{key_prefix}_cost_group")
        
        # Aplicar filtros
        df_filtered = df.copy()
        for column, value in selected_filters.items():
            if value != 'Todas':
                df_filtered = df_filtered[df_filtered[column] == value]
        
        # Mostrar panel de totales
        from src.ui_components import render_totals_panel
        render_totals_panel(df_filtered, f"TOTALES COSTO DE VENTA {title.upper()}")
        
        # Configurar AG-Grid
        gb = AGGridConfigurator.configure_forecast_table(df_filtered)
        
        if selected_filters['BU'] != 'Todas' or not show_grouping:
            gb.configure_column("BU", rowGroup=False, hide=False)
            gb.configure_grid_options(groupDefaultExpanded=0)
        
        # Renderizar AG-Grid
        grid_config = GRID_CONFIGS['forecast_main'].copy()
        grid_config['height'] = AGGridConfigurator.get_grid_height(len(df_filtered), 600)
        
        st.markdown(f"#### 💸 Tabla de Costo de Venta {title}")
        
        AgGrid(
            df_filtered,
            gridOptions=gb.build(),
            height=grid_config['height'],
            theme=grid_config['theme'],
            allow_unsafe_jscode=grid_config['allow_unsafe_jscode'],
            update_mode=grid_config['update_mode'],
            fit_columns_on_grid_load=grid_config['fit_columns_on_grid_load'],
            enable_enterprise_modules=grid_config['enable_enterprise_modules']
        )
        
        # Exportación
        st.markdown("#### 📥 Exportar Datos")
        from src.ui_components import render_export_buttons
        render_export_buttons(df_filtered, f'{key_prefix}_cost', f'{key_prefix}_cost_export')
    
    def generate_consolidated_totals_excel(self, forecast_table: Dict, cost_table: Dict, sheet_prefix: str) -> BytesIO:
        """
        Genera Excel con totales consolidados por Empresa y BU.
        
        Args:
            forecast_table: Dict con datos de forecast
            cost_table: Dict con datos de costo
            sheet_prefix: Prefijo para nombres de hojas
            
        Returns:
            BytesIO con archivo Excel
        """
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        
        df_forecast = pd.DataFrame(forecast_table['data'])
        df_cost = pd.DataFrame(cost_table['data'])
        
        # Excluir filas de totales
        df_forecast = df_forecast[df_forecast['Proyecto'] != 'TOTAL FORECAST'].copy()
        df_cost = df_cost[df_cost['Proyecto'] != 'TOTAL COSTO'].copy()
        
        buffer = BytesIO()
        
        with pd.ExcelWriter(buffer, engine='openpyxl') as writer:
            # Hoja 1: Totales Generales
            totals_data_forecast = []
            totals_data_cost = []
            
            for empresa in df_forecast['Empresa'].unique():
                df_empresa_f = df_forecast[df_forecast['Empresa'] == empresa]
                df_empresa_c = df_cost[df_cost['Empresa'] == empresa]
                
                numeric_cols_f = df_empresa_f.select_dtypes(include=['number']).columns
                numeric_cols_f = [col for col in numeric_cols_f if col not in ['BU']]
                total_f = sum([pd.to_numeric(df_empresa_f[col], errors='coerce').sum() for col in numeric_cols_f])
                
                numeric_cols_c = df_empresa_c.select_dtypes(include=['number']).columns
                numeric_cols_c = [col for col in numeric_cols_c if col not in ['BU']]
                total_c = sum([pd.to_numeric(df_empresa_c[col], errors='coerce').sum() for col in numeric_cols_c])
                
                totals_data_forecast.append({'Empresa': empresa, f'Total Forecast {sheet_prefix}': total_f})
                totals_data_cost.append({'Empresa': empresa, f'Total Costo {sheet_prefix}': total_c})
            
            df_totals_f = pd.DataFrame(totals_data_forecast)
            df_totals_c = pd.DataFrame(totals_data_cost)
            df_totals = pd.merge(df_totals_f, df_totals_c, on='Empresa')
            
            df_totals.to_excel(writer, sheet_name=f'Totales {sheet_prefix}', index=False)
            
            # Formatear hoja de totales
            ws_totals = writer.sheets[f'Totales {sheet_prefix}']
            for row in ws_totals.iter_rows(min_row=2, max_row=len(df_totals)+1, min_col=2, max_col=3):
                for cell in row:
                    cell.number_format = '$#,##0.00'
            
            # Hojas por BU
            for bu in df_forecast['BU'].unique():
                df_bu_f = df_forecast[df_forecast['BU'] == bu]
                df_bu_c = df_cost[df_cost['BU'] == bu]
                
                bu_totals_f = []
                bu_totals_c = []
                
                for empresa in df_bu_f['Empresa'].unique():
                    df_empresa_f = df_bu_f[df_bu_f['Empresa'] == empresa]
                    df_empresa_c = df_bu_c[df_bu_c['Empresa'] == empresa]
                    
                    numeric_cols_f = df_empresa_f.select_dtypes(include=['number']).columns
                    numeric_cols_f = [col for col in numeric_cols_f if col not in ['BU']]
                    total_f = sum([pd.to_numeric(df_empresa_f[col], errors='coerce').sum() for col in numeric_cols_f])
                    
                    numeric_cols_c = df_empresa_c.select_dtypes(include=['number']).columns
                    numeric_cols_c = [col for col in numeric_cols_c if col not in ['BU']]
                    total_c = sum([pd.to_numeric(df_empresa_c[col], errors='coerce').sum() for col in numeric_cols_c])
                    
                    bu_totals_f.append({'Empresa': empresa, 'Total Forecast': total_f})
                    bu_totals_c.append({'Empresa': empresa, 'Total Costo': total_c})
                
                df_bu_totals_f = pd.DataFrame(bu_totals_f)
                df_bu_totals_c = pd.DataFrame(bu_totals_c)
                df_bu_totals = pd.merge(df_bu_totals_f, df_bu_totals_c, on='Empresa')
                
                sheet_name = f'{bu} {sheet_prefix}'
                df_bu_totals.to_excel(writer, sheet_name=sheet_name, index=False)
                
                # Formatear
                ws_bu = writer.sheets[sheet_name]
                for row in ws_bu.iter_rows(min_row=2, max_row=len(df_bu_totals)+1, min_col=2, max_col=3):
                    for cell in row:
                        cell.number_format = '$#,##0.00'
        
        buffer.seek(0)
        return buffer
