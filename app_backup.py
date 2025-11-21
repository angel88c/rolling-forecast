"""
Aplicación principal de Streamlit para Forecast Financiero.

Esta aplicación proporciona una interfaz web intuitiva para generar
proyecciones de ingresos por facturación basadas en el pipeline de oportunidades.
"""

import streamlit as st
import pandas as pd
import plotly.express as px
import plotly.graph_objects as go
from datetime import datetime
from typing import Dict, Any
import sys
import os
import logging
from st_aggrid import AgGrid, GridOptionsBuilder, JsCode

# Agregar el directorio src al path para importar módulos
sys.path.append(os.path.join(os.path.dirname(__file__), 'src'))
sys.path.append(os.path.join(os.path.dirname(__file__), 'config'))

from src.data_processor import DataProcessor
from src.forecast_calculator import ForecastCalculator
from src.validators import DataValidator
from src.exporter import ForecastExporter, ReportGenerator
from src.aggrid_utils import AGGridConfigurator, AGGridExporter, GRID_CONFIGS
from src.grid_utils import GridResponseHandler
from src.kpi_processor import KPIProcessor
from src.ui_styles import apply_custom_styles, create_section_header, format_currency
from src.formatters import (
    format_currency as fmt_currency, 
    format_dataframe_currency_columns,
    format_aggrid_currency_columns,
    format_business_unit_icon,
    format_compact_currency,
    create_summary_stats
)
from config.settings import *


# Configurar logging
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)


class ForecastApp:
    """
    Clase principal de la aplicación Streamlit.
    
    Esta clase organiza toda la lógica de la interfaz de usuario
    y coordina los diferentes módulos de la aplicación.
    """
    
    def __init__(self):
        """Inicializa la aplicación con sus componentes."""
        self.validator = DataValidator()
        self.processor = DataProcessor()
        self.calculator = ForecastCalculator()
        self.exporter = ForecastExporter()
        self.report_generator = ReportGenerator()
        
        # Configurar página
        st.set_page_config(
            page_title=APP_CONFIG.APP_TITLE,
            page_icon="📊",
            layout="wide",
            initial_sidebar_state="expanded"
        )
    
    def run(self):
        """Ejecuta la aplicación principal."""
        # Aplicar estilos personalizados
        apply_custom_styles()
        
        self._render_header()
        self._render_sidebar()
        self._render_main_content()
    
    def _render_header(self):
        """Renderiza el encabezado de la aplicación."""
        # Header principal con diseño mejorado
        st.markdown(f"""
        <div style="text-align: center; padding: 2rem 0; background: linear-gradient(135deg, #1f4e79 0%, #2E86AB 100%); 
                    margin: -1rem -1rem 2rem -1rem; border-radius: 0 0 15px 15px; color: white;">
            <h1 style="margin: 0; font-size: 2.5rem; font-weight: 700; text-shadow: 0 2px 4px rgba(0,0,0,0.3);">
                📊 {SETTINGS['app_config'].APP_TITLE}
            </h1>
            <p style="margin: 0.5rem 0 0 0; font-size: 1.1rem; opacity: 0.9;">
                Proyecciones de ingresos por facturación basadas en pipeline de oportunidades
            </p>
        </div>
        """, unsafe_allow_html=True)
        
        # Mostrar reglas de negocio en un expander
        with st.expander("📋 Reglas de Negocio Aplicadas"):
            col1, col2 = st.columns(2)
            
            with col1:
                st.markdown("**Reglas Generales:**")
                st.markdown(f"- Lead Time mínimo: {BUSINESS_RULES.MIN_LEAD_TIME} semanas")
                st.markdown(f"- Factor de castigo financiero: {BUSINESS_RULES.FINANCIAL_PENALTY_FACTOR_DEFAULT*100}%")
                st.markdown(f"- Días para DR: {BUSINESS_RULES.DR_DAYS_OFFSET}")
                st.markdown(f"- Días para SAT: {BUSINESS_RULES.SAT_DAYS_OFFSET}")
            
            with col2:
                st.markdown("**Porcentajes de Facturación (sin PIA):**")
                st.markdown(f"- INICIO: {BUSINESS_RULES.INICIO_PERCENTAGE*100}%")
                st.markdown(f"- DR: {BUSINESS_RULES.DR_PERCENTAGE*100}%")
                st.markdown(f"- FAT: {BUSINESS_RULES.FAT_PERCENTAGE*100}%")
                st.markdown(f"- SAT: {BUSINESS_RULES.SAT_PERCENTAGE*100}%")
    
    def _render_file_uploader(self, label, key, file_types=['xlsx'], help_text=None):
        """Componente reutilizable para subir archivos."""
        uploaded_file = st.file_uploader(
            label,
            type=file_types,
            help=help_text,
            key=key
        )
        return uploaded_file
    
    def _render_filters_row(self, df, filter_configs):
        """Componente reutilizable para renderizar fila de filtros.
        
        Args:
            df: DataFrame con los datos
            filter_configs: Lista de dicts con configuración de filtros
                [{'column': 'Empresa', 'label': 'Filtrar por Empresa', 'key': 'filter_empresa', 'default': 'Todas'}]
        
        Returns:
            dict: Diccionario con los valores seleccionados
        """
        cols = st.columns(len(filter_configs))
        selected_values = {}
        
        for idx, (col, config) in enumerate(zip(cols, filter_configs)):
            with col:
                column = config['column']
                options = [config.get('default', 'Todas')] + sorted(df[column].dropna().unique().tolist())
                selected = st.selectbox(
                    config['label'],
                    options,
                    key=config['key']
                )
                selected_values[column] = selected
        
        return selected_values
    
    def _render_export_buttons(self, df, filename_prefix, key_prefix):
        """Componente reutilizable para botones de exportación."""
        col1, col2 = st.columns(2)
        
        with col1:
            if st.button("📥 Exportar Excel", key=f"{key_prefix}_excel"):
                buffer = self._export_to_excel_with_format(df, filename_prefix)
                st.download_button(
                    label="⬇️ Descargar Excel",
                    data=buffer.getvalue(),
                    file_name=f"{filename_prefix}_{datetime.now().strftime('%Y%m%d')}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    key=f"{key_prefix}_download_excel"
                )
        
        with col2:
            if st.button("📥 Exportar CSV", key=f"{key_prefix}_csv"):
                csv_data = df.to_csv(index=False)
                st.download_button(
                    label="⬇️ Descargar CSV",
                    data=csv_data,
                    file_name=f"{filename_prefix}_{datetime.now().strftime('%Y%m%d')}.csv",
                    mime="text/csv",
                    key=f"{key_prefix}_download_csv"
                )
    
    def _render_sidebar(self):
        """Renderiza la barra lateral simplificada."""
        
        # Logo de iBtest en la parte superior
        try:
            st.sidebar.image("logo_ibtest.png", width=200)
        except:
            st.sidebar.markdown("### iBtest")
        
        st.sidebar.markdown("---")
        st.sidebar.markdown("**📊 Forecast & KPI Manager**")
        st.sidebar.caption("v3.0 - Gestión de Proyectos")
        
        # Configuración de reglas de negocio editables
        st.sidebar.header("⚙️ Reglas de Negocio")
        
        # Factor de castigo financiero con number_input
        st.session_state.penalty_default = st.sidebar.number_input(
            "Factor Castigo (General) %",
            min_value=10, max_value=100, value=40, step=5,
            help="Factor de castigo para probabilidades diferentes a 60%"
        ) / 100.0
        
        st.session_state.penalty_60 = st.sidebar.number_input(
            "Factor Castigo (60%) %",
            min_value=10, max_value=100, value=60, step=5,
            help="Factor de castigo para probabilidad del 60%"
        ) / 100.0
        
        # Porcentajes de facturación
        with st.sidebar.expander("📊 Porcentajes de Facturación"):
            st.session_state.inicio_pct = st.slider(
                "INICIO (%)", min_value=0, max_value=100, value=30, step=5
            ) / 100
            
            st.session_state.dr_pct = st.slider(
                "DR (%)", min_value=0, max_value=100, value=30, step=5
            ) / 100
            
            st.session_state.fat_pct = st.slider(
                "FAT (%)", min_value=0, max_value=100, value=30, step=5
            ) / 100
            
            st.session_state.sat_pct = st.slider(
                "SAT (%)", min_value=0, max_value=100, value=10, step=5
            ) / 100
            
            # Validar que sumen 100%
            total_pct = (st.session_state.inicio_pct + st.session_state.dr_pct + 
                        st.session_state.fat_pct + st.session_state.sat_pct)
            
            if abs(total_pct - 1.0) > 0.01:
                st.warning(f"⚠️ Los porcentajes suman {total_pct*100:.0f}%, no 100%")
        
        # Botón de procesamiento
        if st.sidebar.button("🚀 Procesar Forecast", type="primary"):
            if hasattr(st.session_state, 'uploaded_file'):
                self._process_forecast()
            else:
                st.sidebar.error("Por favor, sube un archivo primero")
        
        # # Botón para procesar KPIs
        # if st.sidebar.button("📊 Procesar KPIs PM-008", type="secondary"):
        #     if hasattr(st.session_state, 'uploaded_file_kpis'):
        #         self._process_kpis()
        #     else:
        #         st.sidebar.error("Por favor, sube el archivo de KPIs primero")
        
        # Opciones de exportación
        if hasattr(st.session_state, 'forecast_results'):
            st.sidebar.header("📥 Exportar Resultados")
            
            if st.sidebar.button("📊 Descargar Excel"):
                self._export_excel()
            
            if st.sidebar.button("📄 Descargar CSV"):
                self._export_csv()
        
        # Información de la aplicación
        st.sidebar.markdown("---")
        st.sidebar.markdown("**Versión:** 1.0.2")
        st.sidebar.markdown("**Última actualización:** " + datetime.now().strftime("%d/%m/%Y"))
    
    def _render_main_content(self):
        """Renderiza el contenido principal con pestañas."""
        tabs = st.tabs([
            "📊 Forecast", 
            "💰 Costo de Venta",
            "📉 Forecast <60%",
            "💸 Costo Venta <60%",
            "📋 KPIs PM-008", 
            "💵 Costo Venta KPIs", 
            #"📈 Gráficos", 
            #"🎯 Análisis", 
            #"🤖 Chatbot"
        ])
        
        with tabs[0]:
            self._render_forecast_tab()
        
        with tabs[1]:
            self._render_cost_of_sale_tab()
        
        with tabs[2]:
            self._render_forecast_low_prob_tab()
        
        with tabs[3]:
            self._render_cost_of_sale_low_prob_tab()
        
        with tabs[4]:
            self._render_kpi_billing_tab()
        
        with tabs[5]:
            self._render_kpi_cost_tab()
        
        #with tabs[6]:
        #    self._render_charts_tab()
        
        #with tabs[7]:
        #    self._render_analysis_tab()
        
        #with tabs[8]:
        #    self._render_chatbot_tab()
    
    def _render_forecast_tab(self):
        """Pestaña de Forecast Pipeline."""
        if hasattr(st.session_state, 'forecast_results'):
            results = st.session_state.forecast_results
            self._render_forecast_table(results['forecast_table'])
        else:
            self._render_forecast_empty_state()
    
    def _render_forecast_empty_state(self):
        """Estado vacío para pestaña de Forecast."""
        create_section_header("Forecast Pipeline", "Oportunidades >= 60%", "📊")
        
        col_upload, col_process = st.columns([3, 1])
        with col_upload:
            uploaded_file = self._render_file_uploader(
                "📁 Subir archivo C&N Funnel",
                key="forecast_uploader_empty",
                help_text="Archivo Excel con oportunidades del pipeline"
            )
            if uploaded_file:
                st.session_state.uploaded_file = uploaded_file
        
        with col_process:
            if st.button("🔄 Procesar", key="process_forecast_empty", use_container_width=True):
                if hasattr(st.session_state, 'uploaded_file'):
                    self._process_forecast()
                    st.rerun()
                else:
                    st.error("Sube un archivo primero")
        
        st.info("👆 Sube y procesa un archivo Excel para visualizar el forecast del pipeline")
    
    def _render_cost_of_sale_tab(self):
        """Pestaña de Costo de Venta Forecast."""
        if hasattr(st.session_state, 'forecast_results'):
            results = st.session_state.forecast_results
            
            # Verificar si hay datos de costo de venta principal
            if results.get('cost_of_sale_table') is None or len(results['cost_of_sale_table']['data']) == 0:
                st.info("ℹ️ No hay datos de costo de venta principal. Este archivo fue procesado solo para oportunidades < 60%.")
                st.info("👈 Para ver el costo de venta completo, procesa un archivo en la pestaña 'Forecast' o ve a 'Costo Venta <60%'")
                return
            
            self._render_cost_of_sale_table(results['cost_of_sale_table'])
        else:
            st.info("👈 Procesa un forecast en la pestaña 'Forecast' primero")
    
    def _render_forecast_low_prob_tab(self):
        """Pestaña de Forecast para oportunidades con probabilidad < 60%."""
        create_section_header("Forecast Pipeline", "Oportunidades < 60%", "📊")
        
        # File uploader integrado en la pestaña
        col_upload, col_process = st.columns([3, 1])
        with col_upload:
            uploaded_file = self._render_file_uploader(
                "📁 Subir archivo de Forecast (Oportunidades < 60%)",
                key="forecast_low_prob_uploader",
                help_text="Archivo Excel con oportunidades del pipeline (se filtrarán automáticamente las < 60%)"
            )
            if uploaded_file:
                st.session_state.uploaded_file = uploaded_file
        
        with col_process:
            if st.button("🔄 Procesar", key="process_forecast_low_prob", use_container_width=True):
                if hasattr(st.session_state, 'uploaded_file'):
                    self._process_forecast_low_prob()
                else:
                    st.error("Sube un archivo primero")
        
        if not hasattr(st.session_state, 'forecast_results'):
            st.info("👆 Sube y procesa un archivo para visualizar el forecast de oportunidades < 60%")
            return
        
        results = st.session_state.forecast_results
        
        # Verificar si hay datos de oportunidades < 60%
        if results.get('forecast_table_low_prob') is None:
            st.info("ℹ️ No hay oportunidades con probabilidad menor al 60% en el forecast procesado")
            return
        
        # Información sobre el filtro aplicado
        st.info("📉 Esta tabla muestra únicamente oportunidades con **probabilidad < 60%** con factores de castigo aplicados")
        
        # Renderizar la tabla usando el mismo método pero con datos filtrados
        self._render_forecast_table_low_prob(results['forecast_table_low_prob'], results['summary_low_prob'])
    
    def _render_cost_of_sale_low_prob_tab(self):
        """Pestaña de Costo de Venta para oportunidades con probabilidad < 60%."""
        if hasattr(st.session_state, 'forecast_results'):
            results = st.session_state.forecast_results
            
            # Verificar si hay datos de oportunidades < 60%
            if results.get('cost_of_sale_table_low_prob') is None:
                st.info("ℹ️ No hay oportunidades con probabilidad menor al 60% en el forecast procesado")
                return
            
            # Información sobre el filtro aplicado
            st.info("💸 Esta tabla muestra el costo de venta únicamente de oportunidades con **probabilidad < 60%**")
            
            # Renderizar la tabla de costo de venta
            self._render_cost_of_sale_table_low_prob(results['cost_of_sale_table_low_prob'], results['summary_low_prob'])
        else:
            st.info("👈 Procesa un forecast en la pestaña 'Forecast' primero")
    
    def _render_kpi_billing_tab(self):
        """Pestaña de KPIs Billing."""
        if hasattr(st.session_state, 'kpi_results'):
            self._render_kpi_billing_table()
        else:
            self._render_kpi_empty_state()
    
    def _render_kpi_empty_state(self):
        """Estado vacío para pestaña de KPIs."""
        create_section_header("KPIs PM-008", "Billing de proyectos", "📋")
        
        col_upload, col_process = st.columns([3, 1])
        with col_upload:
            uploaded_file_kpi = self._render_file_uploader(
                "📁 Subir archivo KPIs PM-008",
                key="kpi_uploader_empty",
                help_text="Archivo Excel con KPIs PM-008"
            )
            if uploaded_file_kpi:
                st.session_state.uploaded_file_kpis = uploaded_file_kpi
        
        with col_process:
            if st.button("🔄 Procesar", key="process_kpi_empty", use_container_width=True):
                if hasattr(st.session_state, 'uploaded_file_kpis'):
                    self._process_kpis()
                    st.rerun()
                else:
                    st.error("Sube un archivo primero")
        
        st.info("👆 Sube y procesa un archivo Excel de KPIs PM-008")
    
    def _render_kpi_cost_tab(self):
        """Pestaña de Costo de Venta KPIs."""
        if hasattr(st.session_state, 'kpi_results'):
            self._render_kpi_cost_of_sale_table()
        else:
            st.info("👈 Procesa KPIs en la pestaña 'KPIs PM-008' primero")
    
    def _render_charts_tab(self):
        """Pestaña de gráficos."""
        if hasattr(st.session_state, 'forecast_results'):
            results = st.session_state.forecast_results
            self._render_charts(results['summary'], results['billing_events'])
        else:
            st.info("👈 Procesa un forecast primero para visualizar gráficos")
    
    def _render_analysis_tab(self):
        """Pestaña de análisis."""
        if hasattr(st.session_state, 'forecast_results'):
            results = st.session_state.forecast_results
            self._render_analysis(results['billing_events'])
        else:
            st.info("👈 Procesa un forecast primero para ver análisis")
    
    def _render_chatbot_tab(self):
        """Pestaña de chatbot."""
        if hasattr(st.session_state, 'forecast_results'):
            results = st.session_state.forecast_results
            self._render_chatbot(results)
        else:
            st.info("👈 Procesa un forecast primero para usar el chatbot")
    
    def _render_welcome_screen(self):
        """Renderiza la pantalla de bienvenida."""
        st.markdown("## 👋 Bienvenido")
        st.markdown("""
        Para comenzar, sube tu archivo C&N Funnel en la barra lateral y haz clic en **Procesar Forecast**.
        
        ### 📋 Requisitos del archivo:
        - Formato: Excel (.xlsx)
        - Debe contener las columnas requeridas
        
        ### 🔄 Proceso:
        1. **Validación**: Se verifican los datos de entrada
        2. **Limpieza**: Se procesan y ajustan los datos
        3. **Cálculo**: Se aplican las reglas de negocio
        4. **Visualización**: Se muestran los resultados
        """)
        
        # # Mostrar ejemplo de estructura de datos
        # with st.expander("📊 Estructura de Datos Esperada"):
        #     example_data = {
        #         'Opportunity Name': ['Proyecto A', 'Proyecto B', 'Proyecto C'],
        #         'BU': ['FCT', 'ICT', 'IAT'],
        #         'Amount': [100000, 50000, 75000],
        #         'Close Date': ['30/04/2025', '15/05/2025', '01/06/2025'],
        #         'Lead Time': [8, 12, 6],
        #         'Probability (%)  ↑': [0.25, 0.5, 0.25],
        #         'Paid in Advance': [0, 10000, 0],
        #         'Payment Terms': ['NET 30', 'NET 30', 'NET 30']
        #     }
            
        #     st.dataframe(pd.DataFrame(example_data), use_container_width=True)
    
    def _process_forecast(self):
        """Procesa el forecast completo."""
        try:
            with st.spinner("Procesando forecast..."):
                # Paso 1: Validar archivo
                file_validation = self.validator.validate_file(st.session_state.uploaded_file)
                if not file_validation.is_valid:
                    st.error("❌ " + "; ".join(file_validation.errors))
                    return
                
                # Paso 2: Leer archivo con detección automática
                df, parsing_report = self.processor.read_excel_file(st.session_state.uploaded_file)
            
                # Verificar que el parsing fue exitoso
                if not parsing_report.get('parsing_success', False):
                    missing_cols = parsing_report.get('validation_result', {}).get('missing_columns', [])
                    if missing_cols:
                        st.error(f"❌ No se pudieron encontrar las siguientes columnas requeridas: {', '.join(missing_cols)}")
                        st.info("💡 Verifica que el archivo tenga las columnas necesarias con nombres similares a: Opportunity Name, BU, Amount, Close Date, Lead Time, Payment Terms, Probability, Paid in Advance")
                        return
                
                # Paso 3: Procesar datos (incluye limpieza y completado)
                df_clean = self.processor.clean_and_prepare_data(df)

                # Paso 4: Validar datos procesados
                data_validation = self.validator.validate_dataframe(df_clean)
                
                # Mostrar advertencias si las hay (pero no bloquear el procesamiento)
                if data_validation.warnings:
                    for warning in data_validation.warnings[:5]:  # Mostrar solo las primeras 5
                        st.warning("⚠️ " + warning)
                
                # Paso 5: Convertir a objetos Opportunity
                opportunities = self.processor.convert_to_opportunities(df_clean)
                
                # Actualizar reglas de negocio con valores editables
                if hasattr(st.session_state, 'penalty_default'):
                    BUSINESS_RULES.FINANCIAL_PENALTY_FACTOR_DEFAULT = st.session_state.penalty_default
                    BUSINESS_RULES.FINANCIAL_PENALTY_FACTOR_60_PERCENT = st.session_state.penalty_60
                    BUSINESS_RULES.INICIO_PERCENTAGE = st.session_state.inicio_pct
                    BUSINESS_RULES.DR_PERCENTAGE = st.session_state.dr_pct
                    BUSINESS_RULES.FAT_PERCENTAGE = st.session_state.fat_pct
                    BUSINESS_RULES.SAT_PERCENTAGE = st.session_state.sat_pct
                
                # Paso 6: Calcular forecast
                billing_events = self.calculator.calculate_forecast(opportunities)
                
                # Paso 7: Generar resumen y tablas (todas las oportunidades)
                summary = self.calculator.generate_forecast_summary(billing_events)
                forecast_table = self.calculator.create_forecast_table(billing_events)
                cost_of_sale_table = self.calculator.create_cost_of_sale_table(billing_events)
                
                # Paso 8: Separar oportunidades por probabilidad
                # Oportunidades con probabilidad < 60% (< 0.60)
                opportunities_low_prob = [opp for opp in opportunities if opp.probability < 0.60]
                
                if opportunities_low_prob:
                    # Calcular forecast solo para oportunidades < 60%
                    billing_events_low_prob = self.calculator.calculate_forecast(opportunities_low_prob)
                    summary_low_prob = self.calculator.generate_forecast_summary(billing_events_low_prob)
                    forecast_table_low_prob = self.calculator.create_forecast_table(billing_events_low_prob)
                    cost_of_sale_table_low_prob = self.calculator.create_cost_of_sale_table(billing_events_low_prob)
                else:
                    # No hay oportunidades < 60%, crear estructuras vacías
                    billing_events_low_prob = []
                    summary_low_prob = None
                    forecast_table_low_prob = None
                    cost_of_sale_table_low_prob = None
                
                # Guardar resultados en session state
                st.session_state.forecast_results = {
                    'billing_events': billing_events,
                    'summary': summary,
                    'forecast_table': forecast_table,
                    'cost_of_sale_table': cost_of_sale_table,
                    'billing_events_low_prob': billing_events_low_prob,
                    'summary_low_prob': summary_low_prob,
                    'forecast_table_low_prob': forecast_table_low_prob,
                    'cost_of_sale_table_low_prob': cost_of_sale_table_low_prob,
                    'processing_summary': self.processor.get_processing_summary(df, df_clean, parsing_report),
                    'validation_result': data_validation,
                    'parsing_report': parsing_report
                }
                
                #st.write(df)
                st.success("✅ " + INFO_MESSAGES['processing_complete'])
                #st.rerun()
                
        except Exception as e:
            logger.error(f"Error en procesamiento: {str(e)}")
            st.error(f"❌ Error: {str(e)}")
    
    def _process_forecast_low_prob(self):
        """Procesa el forecast SOLO para oportunidades con probabilidad < 60%."""
        try:
            with st.spinner("Procesando forecast de oportunidades < 60%..."):
                # Paso 1: Validar archivo
                file_validation = self.validator.validate_file(st.session_state.uploaded_file)
                if not file_validation.is_valid:
                    st.error("❌ " + "; ".join(file_validation.errors))
                    return
                
                # Paso 2: Leer archivo con detección automática
                df, parsing_report = self.processor.read_excel_file(st.session_state.uploaded_file)
            
                # Verificar que el parsing fue exitoso
                if not parsing_report.get('parsing_success', False):
                    missing_cols = parsing_report.get('validation_result', {}).get('missing_columns', [])
                    if missing_cols:
                        st.error(f"❌ No se pudieron encontrar las siguientes columnas requeridas: {', '.join(missing_cols)}")
                        st.info("💡 Verifica que el archivo tenga las columnas necesarias con nombres similares a: Opportunity Name, BU, Amount, Close Date, Lead Time, Payment Terms, Probability, Paid in Advance")
                        return
                
                # Paso 3: Procesar datos (incluye limpieza y completado)
                df_clean = self.processor.clean_and_prepare_data(df)

                # Paso 4: Validar datos procesados
                data_validation = self.validator.validate_dataframe(df_clean)
                
                # Mostrar advertencias si las hay (pero no bloquear el procesamiento)
                if data_validation.warnings:
                    for warning in data_validation.warnings[:5]:  # Mostrar solo las primeras 5
                        st.warning("⚠️ " + warning)
                
                # Paso 5: Convertir a objetos Opportunity
                opportunities_all = self.processor.convert_to_opportunities(df_clean)
                
                # *** FILTRO: Solo mantener oportunidades con probabilidad < 60% ***
                opportunities_low_prob = [opp for opp in opportunities_all if opp.probability < 0.60]
                
                if not opportunities_low_prob:
                    st.warning("⚠️ No se encontraron oportunidades con probabilidad menor al 60% en el archivo")
                    return
                
                st.info(f"📊 Se encontraron {len(opportunities_low_prob)} oportunidades con probabilidad < 60% de {len(opportunities_all)} totales")
                
                # Actualizar reglas de negocio con valores editables
                if hasattr(st.session_state, 'penalty_default'):
                    BUSINESS_RULES.FINANCIAL_PENALTY_FACTOR_DEFAULT = st.session_state.penalty_default
                    BUSINESS_RULES.FINANCIAL_PENALTY_FACTOR_60_PERCENT = st.session_state.penalty_60
                    BUSINESS_RULES.INICIO_PERCENTAGE = st.session_state.inicio_pct
                    BUSINESS_RULES.DR_PERCENTAGE = st.session_state.dr_pct
                    BUSINESS_RULES.FAT_PERCENTAGE = st.session_state.fat_pct
                    BUSINESS_RULES.SAT_PERCENTAGE = st.session_state.sat_pct
                
                # Paso 6: Calcular forecast SOLO para oportunidades < 60%
                billing_events_low_prob = self.calculator.calculate_forecast(opportunities_low_prob)
                
                # Paso 7: Generar resumen y tablas
                summary_low_prob = self.calculator.generate_forecast_summary(billing_events_low_prob)
                forecast_table_low_prob = self.calculator.create_forecast_table(billing_events_low_prob)
                cost_of_sale_table_low_prob = self.calculator.create_cost_of_sale_table(billing_events_low_prob)
                
                # Guardar resultados en session state
                st.session_state.forecast_results = {
                    'billing_events': [],  # Vacío ya que no procesamos todas
                    'summary': None,
                    'forecast_table': {'data': []},
                    'cost_of_sale_table': {'data': []},
                    'billing_events_low_prob': billing_events_low_prob,
                    'summary_low_prob': summary_low_prob,
                    'forecast_table_low_prob': forecast_table_low_prob,
                    'cost_of_sale_table_low_prob': cost_of_sale_table_low_prob,
                    'processing_summary': self.processor.get_processing_summary(df, df_clean, parsing_report),
                    'validation_result': data_validation,
                    'parsing_report': parsing_report
                }
                
                st.success(f"✅ Forecast procesado: {len(opportunities_low_prob)} oportunidades < 60%")
                
        except Exception as e:
            logger.error(f"Error en procesamiento de forecast <60%: {str(e)}")
            st.error(f"❌ Error: {str(e)}")
    
    def _process_kpis(self):
        """Procesa el archivo de KPIs PM-008."""
        try:
            with st.spinner("Procesando KPIs PM-008..."):
                # Inicializar procesador de KPIs
                kpi_processor = KPIProcessor()
                
                # Procesar archivo
                kpi_results = kpi_processor.process_kpi_file(st.session_state.uploaded_file_kpis)
                
                # Guardar resultados en session state
                st.session_state.kpi_results = kpi_results
                
                st.success(f"✅ KPIs procesados: {kpi_results['filtered_count']} proyectos activos")
                st.rerun()
                
        except Exception as e:
            logger.error(f"Error procesando KPIs: {str(e)}")
            st.error(f"❌ Error al procesar KPIs: {str(e)}")
    
    def _render_key_metrics(self, summary):
        
        
        col1, col2, col3, col4 = st.columns(4)
        
        with col1:
            st.metric(
                "💰 Total Forecast",
                fmt_currency(summary.total_amount, decimals=2),
                help="Monto total proyectado con ajustes aplicados"
            )
        
        with col2:
            st.metric(
                "🎯 Oportunidades",
                f"{summary.total_opportunities:,}",
                help="Número de oportunidades procesadas"
            )
        
        with col3:
            st.metric(
                "📅 Eventos",
                f"{summary.total_events:,}",
                help="Número total de eventos de facturación"
            )
        
        with col4:
            st.metric(
                "⏱️ Duración",
                f"{summary.duration_months} meses",
                help="Duración del forecast en meses"
            )
        
        # Botón de descarga de totales consolidados
        st.markdown("---")
        col_download, col_info = st.columns([1, 3])
        
        with col_download:
            if st.button("📊 Descargar Reporte Consolidado de Totales", key="download_consolidated", help="Descarga Excel con totales por Empresa y BU", use_container_width=True):
                try:
                    excel_buffer = self._generate_consolidated_totals_excel()
                    st.download_button(
                        label="⬇️ Descargar Excel Consolidado",
                        data=excel_buffer.getvalue(),
                        file_name=f"reporte_consolidado_totales_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        key="download_consolidated_btn"
                    )
                except Exception as e:
                    st.error(f"Error al generar reporte: {str(e)}")
        
        with col_info:
            st.info("📋 Incluye: Hoja 'Totales' con resumen general + Una hoja por cada BU (FCT, ICT, IAT, etc.) con sus totales específicos")
    
    def _generate_consolidated_totals_excel(self):
        """Genera un Excel con totales consolidados por Empresa y BU."""
        from io import BytesIO
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        
        output = BytesIO()
        
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            # Obtener datos de todas las fuentes
            forecast_data = pd.DataFrame(st.session_state.forecast_results['forecast_table']['data']) if hasattr(st.session_state, 'forecast_results') else pd.DataFrame()
            kpi_data = pd.DataFrame(st.session_state.kpi_results['data']) if hasattr(st.session_state, 'kpi_results') else pd.DataFrame()
            
            # Identificar columnas de meses (mantienen el formato original de las tablas)
            text_columns = ['Proyecto', 'BU', 'Empresa', 'Company', 'Location', 'Status', 'Customer', '% Facturación', 'Amount Total', 'Gross Margin', 'Costo de Venta', 'Total PO']
            
            # Definir columnas de meses desde los datos originales
            month_cols = []
            month_cols_kpi = []
            
            if not forecast_data.empty:
                month_cols = [col for col in forecast_data.columns if col not in text_columns]
                # Ordenar cronológicamente (más antiguo primero)
                try:
                    month_cols = sorted(month_cols, key=lambda x: pd.to_datetime(x, format='%b %Y'))
                except:
                    # Si el formato no es estándar, intentar parseo general
                    try:
                        month_cols = sorted(month_cols, key=lambda x: pd.to_datetime(x))
                    except:
                        # Si no se puede parsear, dejar en orden original
                        pass
            
            if not kpi_data.empty:
                month_cols_kpi = [col for col in kpi_data.columns if col not in text_columns]
                # Ordenar cronológicamente (más antiguo primero)
                try:
                    month_cols_kpi = sorted(month_cols_kpi, key=lambda x: pd.to_datetime(x, format='%b %Y'))
                except:
                    # Si el formato no es estándar, intentar parseo general
                    try:
                        month_cols_kpi = sorted(month_cols_kpi, key=lambda x: pd.to_datetime(x))
                    except:
                        # Si no se puede parsear, dejar en orden original
                        pass
            
            # === HOJA 1: Totales Generales ===
            totals_rows = []
            
            if not forecast_data.empty and month_cols:
                
                # Forecast Pipeline SAPI
                forecast_sapi = forecast_data[forecast_data['Empresa'] == 'SAPI']
                row_sapi = {'Categoría': 'Forecast Pipeline SAPI'}
                for col in month_cols:
                    row_sapi[col] = pd.to_numeric(forecast_sapi[col], errors='coerce').sum()
                totals_rows.append(row_sapi)
                
                # Forecast Pipeline LLC
                forecast_llc = forecast_data[forecast_data['Empresa'] == 'LLC']
                row_llc = {'Categoría': 'Forecast Pipeline LLC'}
                for col in month_cols:
                    row_llc[col] = pd.to_numeric(forecast_llc[col], errors='coerce').sum()
                totals_rows.append(row_llc)
                
                # Costo de Venta Pipeline SAPI (aparece solo en último mes de facturación por proyecto)
                if 'Costo de Venta' in forecast_data.columns:
                    row_cost_sapi = {'Categoría': 'Costo de Venta Pipeline SAPI'}
                    for col in month_cols:
                        cost = 0
                        for idx in forecast_sapi.index:
                            # Encontrar último mes con facturación para este proyecto
                            last_month = None
                            for m in month_cols:
                                if pd.to_numeric(forecast_sapi.loc[idx, m], errors='coerce') > 0:
                                    last_month = m
                            # Si este es el último mes, agregar el costo
                            if last_month == col:
                                cost += pd.to_numeric(forecast_sapi.loc[idx, 'Costo de Venta'], errors='coerce')
                        row_cost_sapi[col] = cost
                    totals_rows.append(row_cost_sapi)
                    
                    # Costo de Venta Pipeline LLC
                    row_cost_llc = {'Categoría': 'Costo de Venta Pipeline LLC'}
                    for col in month_cols:
                        cost = 0
                        for idx in forecast_llc.index:
                            # Encontrar último mes con facturación para este proyecto
                            last_month = None
                            for m in month_cols:
                                if pd.to_numeric(forecast_llc.loc[idx, m], errors='coerce') > 0:
                                    last_month = m
                            # Si este es el último mes, agregar el costo
                            if last_month == col:
                                cost += pd.to_numeric(forecast_llc.loc[idx, 'Costo de Venta'], errors='coerce')
                        row_cost_llc[col] = cost
                    totals_rows.append(row_cost_llc)
            
            if not kpi_data.empty and month_cols_kpi:
                
                # Forecast KPIs SAPI
                kpi_sapi = kpi_data[kpi_data['Location'] == 'SAPI']
                row_kpi_sapi = {'Categoría': 'Forecast KPIs SAPI'}
                for col in month_cols_kpi:
                    row_kpi_sapi[col] = pd.to_numeric(kpi_sapi[col], errors='coerce').sum()
                totals_rows.append(row_kpi_sapi)
                
                # Forecast KPIs LLC
                kpi_llc = kpi_data[kpi_data['Location'] == 'LLC']
                row_kpi_llc = {'Categoría': 'Forecast KPIs LLC'}
                for col in month_cols_kpi:
                    row_kpi_llc[col] = pd.to_numeric(kpi_llc[col], errors='coerce').sum()
                totals_rows.append(row_kpi_llc)
                
                # Costo de Venta KPIs SAPI (aparece solo en último mes de facturación por proyecto)
                if 'Costo de Venta' in kpi_data.columns:
                    row_kpi_cost_sapi = {'Categoría': 'Costo de Venta KPIs SAPI'}
                    for col in month_cols_kpi:
                        cost = 0
                        for idx in kpi_sapi.index:
                            # Encontrar último mes con facturación para este proyecto
                            last_month = None
                            for m in month_cols_kpi:
                                if pd.to_numeric(kpi_sapi.loc[idx, m], errors='coerce') > 0:
                                    last_month = m
                            # Si este es el último mes, agregar el costo
                            if last_month == col:
                                cost += pd.to_numeric(kpi_sapi.loc[idx, 'Costo de Venta'], errors='coerce')
                        row_kpi_cost_sapi[col] = cost
                    totals_rows.append(row_kpi_cost_sapi)
                    
                    # Costo de Venta KPIs LLC
                    row_kpi_cost_llc = {'Categoría': 'Costo de Venta KPIs LLC'}
                    for col in month_cols_kpi:
                        cost = 0
                        for idx in kpi_llc.index:
                            # Encontrar último mes con facturación para este proyecto
                            last_month = None
                            for m in month_cols_kpi:
                                if pd.to_numeric(kpi_llc.loc[idx, m], errors='coerce') > 0:
                                    last_month = m
                            # Si este es el último mes, agregar el costo
                            if last_month == col:
                                cost += pd.to_numeric(kpi_llc.loc[idx, 'Costo de Venta'], errors='coerce')
                        row_kpi_cost_llc[col] = cost
                    totals_rows.append(row_kpi_cost_llc)
            
            # Crear DataFrame de totales
            df_totals = pd.DataFrame(totals_rows)
            
            # Reordenar columnas: 'Categoría' primero, luego meses en orden cronológico
            if not df_totals.empty:
                # Obtener todas las columnas de meses (todas excepto 'Categoría')
                all_month_cols = [col for col in df_totals.columns if col != 'Categoría']
                # Ordenar cronológicamente
                try:
                    all_month_cols_sorted = sorted(all_month_cols, key=lambda x: pd.to_datetime(x, format='%b %Y'))
                except:
                    try:
                        all_month_cols_sorted = sorted(all_month_cols, key=lambda x: pd.to_datetime(x))
                    except:
                        all_month_cols_sorted = all_month_cols
                # Reordenar: 'Categoría' + meses ordenados
                df_totals = df_totals[['Categoría'] + all_month_cols_sorted]
            
            # Escribir hoja de Totales
            df_totals.to_excel(writer, sheet_name='Totales', index=False)
            
            # Aplicar formato
            worksheet = writer.sheets['Totales']
            self._apply_excel_formatting(worksheet, df_totals)
            
            # === HOJAS POR BU: Una hoja por cada BU ===
            if not forecast_data.empty and 'BU' in forecast_data.columns:
                # Obtener todas las BUs únicas de ambas fuentes
                bus_forecast = set(forecast_data['BU'].dropna().unique())
                bus_kpi = set(kpi_data['BU'].dropna().unique()) if not kpi_data.empty and 'BU' in kpi_data.columns else set()
                all_bus = sorted(bus_forecast | bus_kpi)
                
                # Crear una hoja por cada BU
                for bu in all_bus:
                    if pd.isna(bu):
                        continue
                    
                    bu_rows = []
                    
                    # Datos de Forecast Pipeline para esta BU
                    forecast_bu_sapi = forecast_data[(forecast_data['BU'] == bu) & (forecast_data['Empresa'] == 'SAPI')]
                    forecast_bu_llc = forecast_data[(forecast_data['BU'] == bu) & (forecast_data['Empresa'] == 'LLC')]
                    
                    # Forecast Pipeline SAPI
                    if not forecast_bu_sapi.empty:
                        row_bu_sapi = {'Categoría': 'Forecast Pipeline SAPI'}
                        for col in month_cols:
                            row_bu_sapi[col] = pd.to_numeric(forecast_bu_sapi[col], errors='coerce').sum()
                        bu_rows.append(row_bu_sapi)
                    
                    # Forecast Pipeline LLC
                    if not forecast_bu_llc.empty:
                        row_bu_llc = {'Categoría': 'Forecast Pipeline LLC'}
                        for col in month_cols:
                            row_bu_llc[col] = pd.to_numeric(forecast_bu_llc[col], errors='coerce').sum()
                        bu_rows.append(row_bu_llc)
                    
                    # Costo de Venta Pipeline (aparece solo en último mes de facturación por proyecto)
                    if 'Costo de Venta' in forecast_data.columns:
                        if not forecast_bu_sapi.empty:
                            row_bu_cost_sapi = {'Categoría': 'Costo de Venta Pipeline SAPI'}
                            for col in month_cols:
                                cost = 0
                                for idx in forecast_bu_sapi.index:
                                    # Encontrar último mes con facturación para este proyecto
                                    last_month = None
                                    for m in month_cols:
                                        if pd.to_numeric(forecast_bu_sapi.loc[idx, m], errors='coerce') > 0:
                                            last_month = m
                                    # Si este es el último mes, agregar el costo
                                    if last_month == col:
                                        cost += pd.to_numeric(forecast_bu_sapi.loc[idx, 'Costo de Venta'], errors='coerce')
                                row_bu_cost_sapi[col] = cost
                            bu_rows.append(row_bu_cost_sapi)
                        
                        if not forecast_bu_llc.empty:
                            row_bu_cost_llc = {'Categoría': 'Costo de Venta Pipeline LLC'}
                            for col in month_cols:
                                cost = 0
                                for idx in forecast_bu_llc.index:
                                    # Encontrar último mes con facturación para este proyecto
                                    last_month = None
                                    for m in month_cols:
                                        if pd.to_numeric(forecast_bu_llc.loc[idx, m], errors='coerce') > 0:
                                            last_month = m
                                    # Si este es el último mes, agregar el costo
                                    if last_month == col:
                                        cost += pd.to_numeric(forecast_bu_llc.loc[idx, 'Costo de Venta'], errors='coerce')
                                row_bu_cost_llc[col] = cost
                            bu_rows.append(row_bu_cost_llc)
                    
                    # Datos de KPIs para esta BU
                    if not kpi_data.empty and 'BU' in kpi_data.columns:
                        kpi_bu_sapi = kpi_data[(kpi_data['BU'] == bu) & (kpi_data['Location'] == 'SAPI')]
                        kpi_bu_llc = kpi_data[(kpi_data['BU'] == bu) & (kpi_data['Location'] == 'LLC')]
                        
                        # KPIs SAPI
                        if not kpi_bu_sapi.empty:
                            row_kpi_bu_sapi = {'Categoría': 'Forecast KPIs SAPI'}
                            for col in month_cols_kpi:
                                row_kpi_bu_sapi[col] = pd.to_numeric(kpi_bu_sapi[col], errors='coerce').sum()
                            bu_rows.append(row_kpi_bu_sapi)
                        
                        # KPIs LLC
                        if not kpi_bu_llc.empty:
                            row_kpi_bu_llc = {'Categoría': 'Forecast KPIs LLC'}
                            for col in month_cols_kpi:
                                row_kpi_bu_llc[col] = pd.to_numeric(kpi_bu_llc[col], errors='coerce').sum()
                            bu_rows.append(row_kpi_bu_llc)
                        
                        # Costo de Venta KPIs
                        if 'Costo de Venta' in kpi_data.columns:
                            if not kpi_bu_sapi.empty:
                                row_kpi_cost_sapi = {'Categoría': 'Costo de Venta KPIs SAPI'}
                                for col in month_cols_kpi:
                                    cost = 0
                                    for idx in kpi_bu_sapi.index:
                                        # Encontrar último mes con facturación para este proyecto
                                        last_month = None
                                        for m in month_cols_kpi:
                                            if pd.to_numeric(kpi_bu_sapi.loc[idx, m], errors='coerce') > 0:
                                                last_month = m
                                        # Si este es el último mes, agregar el costo
                                        if last_month == col:
                                            cost += pd.to_numeric(kpi_bu_sapi.loc[idx, 'Costo de Venta'], errors='coerce')
                                    row_kpi_cost_sapi[col] = cost
                                bu_rows.append(row_kpi_cost_sapi)
                            
                            if not kpi_bu_llc.empty:
                                row_kpi_cost_llc = {'Categoría': 'Costo de Venta KPIs LLC'}
                                for col in month_cols_kpi:
                                    cost = 0
                                    for idx in kpi_bu_llc.index:
                                        # Encontrar último mes con facturación para este proyecto
                                        last_month = None
                                        for m in month_cols_kpi:
                                            if pd.to_numeric(kpi_bu_llc.loc[idx, m], errors='coerce') > 0:
                                                last_month = m
                                        # Si este es el último mes, agregar el costo
                                        if last_month == col:
                                            cost += pd.to_numeric(kpi_bu_llc.loc[idx, 'Costo de Venta'], errors='coerce')
                                    row_kpi_cost_llc[col] = cost
                                bu_rows.append(row_kpi_cost_llc)
                    
                    # Crear DataFrame para esta BU y escribirlo en su hoja
                    if bu_rows:
                        df_bu = pd.DataFrame(bu_rows)
                        
                        # Reordenar columnas: 'Categoría' primero, luego meses en orden cronológico
                        if not df_bu.empty:
                            # Obtener todas las columnas de meses (todas excepto 'Categoría')
                            bu_month_cols = [col for col in df_bu.columns if col != 'Categoría']
                            # Ordenar cronológicamente
                            try:
                                bu_month_cols_sorted = sorted(bu_month_cols, key=lambda x: pd.to_datetime(x, format='%b %Y'))
                            except:
                                try:
                                    bu_month_cols_sorted = sorted(bu_month_cols, key=lambda x: pd.to_datetime(x))
                                except:
                                    bu_month_cols_sorted = bu_month_cols
                            # Reordenar: 'Categoría' + meses ordenados
                            df_bu = df_bu[['Categoría'] + bu_month_cols_sorted]
                        
                        # Escribir hoja con el nombre de la BU
                        df_bu.to_excel(writer, sheet_name=str(bu), index=False)
                        
                        # Aplicar formato
                        worksheet_bu = writer.sheets[str(bu)]
                        self._apply_excel_formatting(worksheet_bu, df_bu)
        
        output.seek(0)
        return output
    
    def _apply_excel_formatting(self, worksheet, df):
        """Aplica formato profesional al Excel."""
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        
        # Estilo para encabezados
        header_fill = PatternFill(start_color='2E86AB', end_color='2E86AB', fill_type='solid')
        header_font = Font(bold=True, color='FFFFFF', size=11)
        
        # Aplicar estilo a encabezados
        for cell in worksheet[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Formato de moneda para columnas numéricas
        for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row):
            for idx, cell in enumerate(row):
                if idx > 0 or (idx > 1 and 'BU' in df.columns):  # Columnas numéricas
                    try:
                        if cell.value and isinstance(cell.value, (int, float)):
                            cell.number_format = '$#,##0.00'
                            cell.alignment = Alignment(horizontal='right')
                    except:
                        pass
        
        # Ajustar ancho de columnas
        for column in worksheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 20)
            worksheet.column_dimensions[column_letter].width = adjusted_width
    
    def _render_totals_panel(self, df, label, color='#E3F2FD'):
        """Renderiza un panel de totales separado que siempre es visible."""
        # Calcular totales solo de columnas numéricas
        text_columns = ['Proyecto', 'BU', 'Empresa', 'Company', 'Location', 'Status', 'Customer', '% Facturación']
        
        totals = {}
        for col in df.columns:
            if col not in text_columns:
                try:
                    total = pd.to_numeric(df[col], errors='coerce').sum()
                    if total != 0:  # Solo mostrar si hay valor
                        totals[col] = total
                except:
                    pass
        
        # Crear string para copiar al portapapeles
        totals_text = f"{label}\n" + "\n".join([f"{k}, ${v:,.2f}" for k, v in totals.items()])
        
        # Mostrar panel de totales
        st.markdown(f"### 🧮 {label}")
        
        # Crear DataFrame con los totales en formato tabla
        if totals:
            # Crear DataFrame con una fila de totales
            totals_df = pd.DataFrame([totals])
            
            # Formatear valores como moneda para visualización
            totals_df_formatted = totals_df.copy()
            for col in totals_df_formatted.columns:
                totals_df_formatted[col] = totals_df_formatted[col].apply(lambda x: f"${x:,.2f}")
            
            # Mostrar tabla con st.data_editor
            st.data_editor(
                totals_df_formatted,
                use_container_width=True,
                hide_index=True,
                disabled=True,  # Solo lectura
                column_config={
                    col: st.column_config.TextColumn(
                        col,
                        width="medium",
                    ) for col in totals_df_formatted.columns
                }
            )
            
            # Botón para copiar al portapapeles
            # col_copy, col_space = st.columns([1, 3])
            # with col_copy:
            #     if st.button("📋 Copiar Totales", key=f"copy_{label.replace(' ', '_')}"):
            #         st.code(totals_text, language=None)
            #         st.success("✅ Copiado!")
        else:
            st.info("No hay totales para mostrar")
        
        st.markdown("---")
        return totals
    
    def _render_forecast_table(self, forecast_table):
        """Renderiza la tabla principal del forecast."""
        #create_section_header("Forecast Pipeline", "Oportunidades >= 60%", "📊")
        
        # File uploader integrado en la pestaña
        col_upload, col_process = st.columns([3, 1])
        with col_upload:
            uploaded_file = self._render_file_uploader(
                "📁 Subir archivo de Forecast de Oportunidades",
                key="forecast_uploader",
                help_text="Archivo Excel con oportunidades del pipeline"
            )
            if uploaded_file:
                st.session_state.uploaded_file = uploaded_file
        
        with col_process:
            if st.button("🔄 Procesar", key="process_forecast", use_container_width=True):
                if hasattr(st.session_state, 'uploaded_file'):
                    self._process_forecast()
                else:
                    st.error("Sube un archivo primero")
        
        if not hasattr(st.session_state, 'forecast_results'):
            st.info("👆 Sube y procesa un archivo para visualizar el forecast")
            return
        
        results = st.session_state.forecast_results
        
        # Verificar si hay datos del forecast principal
        if results.get('summary') is None or len(forecast_table['data']) == 0:
            st.info("ℹ️ No hay datos de forecast principal. Este archivo fue procesado solo para oportunidades < 60%.")
            st.info("👈 Para ver el forecast completo, procesa un archivo en esta pestaña o ve a la pestaña 'Forecast <60%'")
            return
        
        # Métricas principales
        self._render_key_metrics(results['summary'])
        
        st.markdown("---")
        
        df = pd.DataFrame(forecast_table['data'])
        
        # Filtros usando componente reutilizable
        filter_configs = [
            {'column': 'Empresa', 'label': '🏢 Empresa', 'key': 'forecast_empresa'},
            {'column': 'BU', 'label': '📋 BU', 'key': 'forecast_bu'}
        ]
        
        col_filters, col_controls = st.columns([3, 1])
        
        with col_filters:
            selected_filters = self._render_filters_row(df, filter_configs)
        
        with col_controls:
            show_grouping = st.checkbox("Agrupar por BU", value=True, key="forecast_group")
        
        # Aplicar filtros
        df_filtered = df.copy()
        
        for column, value in selected_filters.items():
            if value != 'Todas':
                df_filtered = df_filtered[df_filtered[column] == value]
        
        # Mostrar panel de totales separado (siempre visible)
        self._render_totals_panel(df_filtered, "TOTALES FORECAST")
        
        # Configurar AG-Grid (SIN fila de totales en la tabla)
        gb = AGGridConfigurator.configure_forecast_table(df_filtered)
        
        # Configuraciones específicas según filtros
        if selected_filters['BU'] != 'Todas' or not show_grouping:
            # Si hay filtro específico o no se quiere agrupación, desactivar agrupación
            gb.configure_column("BU", rowGroup=False, hide=False)
            gb.configure_grid_options(groupDefaultExpanded=0)
        
        # Renderizar AG-Grid
        grid_config = GRID_CONFIGS['forecast_main'].copy()
        grid_config['height'] = AGGridConfigurator.get_grid_height(len(df_filtered), 600)
        
        st.markdown("#### 📊 Tabla de Forecast")
        
        grid_response = AgGrid(
            df_filtered,
            gridOptions=gb.build(),
            height=grid_config['height'],
            theme=grid_config['theme'],
            allow_unsafe_jscode=grid_config['allow_unsafe_jscode'],
            update_mode=grid_config['update_mode'],
            fit_columns_on_grid_load=grid_config['fit_columns_on_grid_load'],
            enable_enterprise_modules=grid_config['enable_enterprise_modules']
        )
        
        # Mostrar estadísticas y controles adicionales usando GridResponseHandler
        grid_handler = GridResponseHandler(grid_response)
        
        if grid_handler.has_data:
            # Ya no necesitamos filtrar la fila de totales porque no existe
            df_forecast_metrics = grid_handler.data_df.copy()
            
            col_stats1, col_stats2, col_stats3 = st.columns(3)
            
            with col_stats1:
                st.metric("📋 Proyectos", len(df_forecast_metrics))
            
            with col_stats2:
                numeric_columns = df_forecast_metrics.select_dtypes(include=['number']).columns
                numeric_columns = [col for col in numeric_columns if col not in ['BU']]
                if len(numeric_columns) > 0:
                    total_forecast = 0
                    for col in numeric_columns:
                        total_forecast += pd.to_numeric(df_forecast_metrics[col], errors='coerce').sum()
                    st.metric("💰 Total Forecast", f"${total_forecast:,.0f}")
            
            with col_stats3:
                unique_bus = df_forecast_metrics['BU'].nunique()
                st.metric("🏢 BUs Activas", unique_bus)
            
            # Exportación simplificada
            st.markdown("#### 📥 Exportar Datos")
            self._render_export_buttons(df_filtered, 'forecast', 'forecast_export')
    

    def _export_to_excel_with_format(self, df, sheet_name='Datos'):
        """
        Exporta DataFrame a Excel con formato de moneda en columnas numéricas.
        
        Args:
            df: DataFrame a exportar
            sheet_name: Nombre de la hoja
            
        Returns:
            BytesIO: Buffer con archivo Excel formateado
        """
        from io import BytesIO
        from openpyxl.styles import numbers
        
        buffer = BytesIO()
        
        # Identificar columnas numéricas (excluyendo columnas de texto)
        text_columns = ['Proyecto', 'BU', 'Empresa', 'Company', 'Location', 'Status', 'Customer', '% Facturación']
        numeric_columns = [col for col in df.columns if col not in text_columns]
        
        with pd.ExcelWriter(buffer, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name=sheet_name)
            
            # Obtener la hoja de trabajo
            worksheet = writer.sheets[sheet_name]
            
            # Aplicar formato de moneda a columnas numéricas
            for idx, col in enumerate(df.columns, start=1):
                if col in numeric_columns:
                    col_letter = worksheet.cell(row=1, column=idx).column_letter
                    
                    # Aplicar formato de moneda (ej: $1,234.56)
                    for row in range(2, len(df) + 2):  # +2 porque empezamos en row 2 (después del header)
                        cell = worksheet[f'{col_letter}{row}']
                        try:
                            if cell.value and isinstance(cell.value, (int, float)):
                                cell.number_format = '$#,##0.00'
                        except:
                            pass
            
            # Ajustar ancho de columnas
            for column in worksheet.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                worksheet.column_dimensions[column_letter].width = adjusted_width
        
        buffer.seek(0)
        return buffer
    
    def _add_totals_row(self, df, label='TOTAL'):
        """
        Agrega una fila de totales al final del DataFrame.
        
        Args:
            df: DataFrame al que agregar totales
            label: Etiqueta para la fila de totales
            
        Returns:
            DataFrame con fila de totales agregada
        """
        if df.empty:
            return df
        
        # Crear copia del DataFrame
        df_copy = df.copy()
        
        # Identificar columnas no numéricas (textuales)
        text_columns = ['Proyecto', 'BU', 'Empresa', 'Company']
        # Columnas que deben sumarse aunque tengan nombres específicos
        numeric_summary_columns = ['Amount Total', 'Gross Margin', 'Costo de Venta']
        
        # Crear fila de totales
        totals_row = {}
        
        # Primera columna lleva el label
        first_col = df_copy.columns[0]
        totals_row[first_col] = label
        
        # Procesar cada columna
        for col in df_copy.columns:
            if col == first_col:
                continue
            elif col in text_columns:
                # Columnas de texto quedan vacías
                totals_row[col] = ''
            else:
                # Sumar columnas numéricas (incluyendo Amount Total, Gross Margin, Costo de Venta)
                try:
                    # Asegurar que los valores sean numéricos
                    col_values = pd.to_numeric(df_copy[col], errors='coerce').fillna(0)
                    totals_row[col] = col_values.sum()
                except:
                    totals_row[col] = ''
        
        # Convertir a DataFrame y concatenar
        totals_df = pd.DataFrame([totals_row])
        result_df = pd.concat([df_copy, totals_df], ignore_index=True)
        
        return result_df
    
    def _render_forecast_monthly_totals(self, forecast_table, df_filtered):
        """Renderiza los totales mensuales de la tabla de Forecast."""
        if 'monthly_totals' not in forecast_table:
            return
        
        st.markdown("---")
        st.markdown("#### 📊 Totales Mensuales - Forecast")
        
        monthly_totals = forecast_table['monthly_totals']
        
        # Calcular totales solo de los meses que aparecen en el DataFrame filtrado
        month_columns = [col for col in df_filtered.columns if col not in ['Proyecto', 'BU', 'Empresa']]
        
        # Crear DataFrame de totales
        totals_data = []
        for month in month_columns:
            # Sumar valores del DataFrame filtrado
            month_total = df_filtered[month].sum() if month in df_filtered.columns else 0
            totals_data.append({
                'Mes': month,
                'Total': month_total
            })
        
        if totals_data:
            df_totals = pd.DataFrame(totals_data)
            
            # Crear gráfico de barras horizontal con totales
            fig_totals = px.bar(
                df_totals,
                x='Total',
                y='Mes',
                orientation='h',
                title='Totales Mensuales del Forecast',
                labels={'Total': 'Monto Total ($)', 'Mes': 'Período'},
                text='Total'
            )
            
            fig_totals.update_traces(
                texttemplate='$%{text:,.0f}',
                textposition='outside',
                marker_color='#2E86AB',
                textfont_size=10
            )
            
            fig_totals.update_layout(
                height=max(300, len(df_totals) * 35),
                title_font_size=16,
                title_font_color='#1f4e79',
                font=dict(family="Arial, sans-serif", size=11, color="#2c3e50"),
                plot_bgcolor='rgba(0,0,0,0)',
                paper_bgcolor='rgba(0,0,0,0)',
                showlegend=False,
                xaxis_title="Total Facturado ($)",
                yaxis_title="",
                yaxis={'categoryorder':'total ascending'}
            )
            
            st.plotly_chart(fig_totals, use_container_width=True)
            
            # Mostrar tabla con totales
            st.markdown("##### 📋 Tabla de Totales")
            df_totals['Total'] = df_totals['Total'].apply(lambda x: f"${x:,.2f}")
            st.dataframe(
                df_totals,
                use_container_width=True,
                hide_index=True
            )
            
            # Mostrar gran total
            grand_total = sum([item['Total'] for item in totals_data])
            st.metric("💰 Gran Total", f"${grand_total:,.2f}")
    
    def _render_cost_of_sale_table(self, cost_table):
        """Renderiza la tabla completa de Costo de Venta."""
        if len(cost_table['data']) == 0:
            st.warning("No hay datos de costo de venta disponibles")
            return
        
        create_section_header("Costo de Venta", "Costo de venta calculado en el último mes de facturación", "💰")
        
        df_cost = pd.DataFrame(cost_table['data'])
        
        # Controles superiores
        # Filtros de la tabla de costos
        #st.markdown("#### 🔍 Filtros")
        col_ctrl1, col_ctrl2, col_ctrl3, col_ctrl4 = st.columns([2, 2, 1, 1])
        
        with col_ctrl1:
            company_options_cost = ['Todas'] + sorted(df_cost['Empresa'].dropna().unique().tolist())
            selected_company_cost = st.selectbox("Filtrar por Empresa:", company_options_cost, key="cost_company_filter")
        
        with col_ctrl2:
            bu_options_cost = ['Todas'] + sorted(df_cost['BU'].dropna().unique().tolist())
            selected_bu_cost = st.selectbox("Filtrar por BU:", bu_options_cost, key="cost_bu_filter")
        
        with col_ctrl3:
            show_grouping_cost = st.checkbox("Agrupar por BU", value=True, key="group_by_bu_cost")
        
        with col_ctrl4:
            export_format_cost = st.selectbox(
                "Formato de Exportación",
                ["Excel", "CSV"],
                key="export_format_cost"
            )
        
        # Aplicar filtros
        df_cost_filtered = df_cost.copy()
        
        if selected_company_cost != 'Todas':
            df_cost_filtered = df_cost_filtered[df_cost_filtered['Empresa'] == selected_company_cost]
        
        if selected_bu_cost != 'Todas':
            df_cost_filtered = df_cost_filtered[df_cost_filtered['BU'] == selected_bu_cost]
        
        # Mostrar panel de totales separado (siempre visible)
        self._render_totals_panel(df_cost_filtered, "TOTALES COSTO DE VENTA", color='#E8F5E9')
        
        # Configurar AG-Grid con estilo especial para costo de venta (SIN fila de totales)
        gb_cost = AGGridConfigurator.configure_forecast_table(df_cost_filtered)
        
        # Configuraciones específicas según agrupación
        if selected_bu_cost != 'Todas' or not show_grouping_cost:
            # Si hay filtro específico o no se quiere agrupación, desactivar agrupación
            gb_cost.configure_column("BU", rowGroup=False, hide=False)
            gb_cost.configure_grid_options(groupDefaultExpanded=0, getRowStyle=None)
        else:
            # Mantener agrupación por BU
            gb_cost.configure_grid_options(groupDefaultExpanded=1)
        
        # Aplicar color naranja #FCB72F a las columnas de meses con valores > 0
        month_columns = [col for col in df_cost_filtered.columns if col not in ['Proyecto', 'BU', 'Empresa', 'Amount Total', 'Gross Margin', 'Costo de Venta']]
        
        for month_col in month_columns:
            # JsCode para colorear celdas con valores > 0
            cell_style_jscode = JsCode("""
            function(params) {
                if (params.value > 0) {
                    return {
                        'backgroundColor': '#FCB72F',
                        'color': '#000000',
                        'fontWeight': 'bold'
                    }
                }
                return null;
            }
            """)
            
            gb_cost.configure_column(
                month_col,
                cellStyle=cell_style_jscode
            )
        
        # Renderizar AG-Grid
        grid_config_cost = GRID_CONFIGS['forecast_main'].copy()
        grid_config_cost['height'] = AGGridConfigurator.get_grid_height(len(df_cost_filtered), 500)
        
        st.markdown("#### 💰 Tabla de Costo de Venta")
        st.info("📌 El costo de venta se muestra en el **mes del último evento** de cada proyecto (naranja). Los totales se muestran en el panel verde arriba.")
        
        grid_response_cost = AgGrid(
            df_cost_filtered,
            gridOptions=gb_cost.build(),
            height=grid_config_cost['height'],
            theme=grid_config_cost['theme'],
            allow_unsafe_jscode=grid_config_cost['allow_unsafe_jscode'],
            update_mode=grid_config_cost['update_mode'],
            fit_columns_on_grid_load=grid_config_cost['fit_columns_on_grid_load'],
            enable_enterprise_modules=grid_config_cost['enable_enterprise_modules']
        )
        
        # Estadísticas de la tabla de costos
        cost_handler = GridResponseHandler(grid_response_cost)
        
        if cost_handler.has_data:
            # Ya no hay fila de totales, usar datos directamente
            df_cost_metrics = cost_handler.data_df.copy()
            
            col_cost1, col_cost2, col_cost3 = st.columns(3)
            
            with col_cost1:
                st.metric("📋 Proyectos", len(df_cost_metrics))
            
            with col_cost2:
                if 'Costo de Venta' in df_cost_metrics.columns:
                    total_cost = pd.to_numeric(df_cost_metrics['Costo de Venta'], errors='coerce').sum()
                    st.metric("💰 Total Costo de Venta", f"${total_cost:,.0f}")
            
            with col_cost3:
                if 'Gross Margin' in df_cost_metrics.columns:
                    total_margin = pd.to_numeric(df_cost_metrics['Gross Margin'], errors='coerce').sum()
                    st.metric("📈 Total Gross Margin", f"${total_margin:,.0f}")
            
            # Controles de exportación
            st.markdown("#### 🔧 Controles Avanzados")
            col_control1, col_control2, col_control3 = st.columns(3)
            
            with col_control1:
                if st.button("📥 Exportar Datos Visibles", key="export_visible_cost"):
                    try:
                        # Ya no hay fila de totales, exportar directamente
                        df_export_cost = df_cost_filtered.copy()
                        
                        if export_format_cost == "CSV":
                            export_data = df_export_cost.to_csv(index=False)
                            st.download_button(
                                label="⬇️ Descargar CSV",
                                data=export_data,
                                file_name=f"costo_venta_{selected_bu_cost}_{datetime.now().strftime('%Y%m%d')}.csv",
                                mime="text/csv"
                            )
                        else:
                            # Usar función con formato de moneda
                            buffer = self._export_to_excel_with_format(df_export_cost, 'Costo de Venta')
                            st.download_button(
                                label="⬇️ Descargar Excel",
                                data=buffer.getvalue(),
                                file_name=f"costo_venta_{selected_bu_cost}_{datetime.now().strftime('%Y%m%d')}.xlsx",
                                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                            )
                    except Exception as e:
                        st.error(f"Error al exportar: {str(e)}")
            
            with col_control2:
                if cost_handler.has_selection:
                    st.success(f"✅ {cost_handler.selected_count} filas seleccionadas")
                    
                    if st.button("📥 Exportar Seleccionadas", key="export_selected_cost"):
                        try:
                            # Ya no hay fila de totales
                            selected_data = cost_handler.selected_data
                            
                            if export_format_cost == "CSV":
                                csv_data = selected_data.to_csv(index=False)
                                st.download_button(
                                    label="⬇️ Descargar Selección CSV",
                                    data=csv_data,
                                    file_name=f"costo_venta_seleccion_{datetime.now().strftime('%Y%m%d')}.csv",
                                    mime="text/csv"
                                )
                            else:
                                # Usar función con formato de moneda
                                buffer = self._export_to_excel_with_format(selected_data, 'Selección')
                                st.download_button(
                                    label="⬇️ Descargar Selección Excel",
                                    data=buffer.getvalue(),
                                    file_name=f"costo_venta_seleccion_{datetime.now().strftime('%Y%m%d')}.xlsx",
                                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                                )
                        except Exception as e:
                            st.error(f"Error al exportar selección: {str(e)}")
                else:
                    st.info("💡 Selecciona filas para exportar")
            
            with col_control3:
                # Mostrar información de filtros aplicados
                filters_info = {
                    'Empresa': selected_company_cost if selected_company_cost != 'Todas' else None,
                    'BU': selected_bu_cost if selected_bu_cost != 'Todas' else None,
                    'Agrupación': 'Activa' if show_grouping_cost else 'Desactivada'
                }
                active_filters = [f"{k}: {v}" for k, v in filters_info.items() if v]
                
                if active_filters:
                    st.info(f"🔍 {', '.join(active_filters)}")
                else:
                    st.info("📊 Vista completa")
    
    def _render_cost_monthly_totals(self, cost_table, df_cost_filtered):
        """Renderiza los totales mensuales de la tabla de Costo de Venta."""
        if 'monthly_totals' not in cost_table:
            return
        
        st.markdown("---")
        st.markdown("#### 💰 Totales Mensuales - Costo de Venta")
        
        # Calcular totales solo de los meses que aparecen en el DataFrame filtrado
        month_columns = [col for col in df_cost_filtered.columns 
                        if col not in ['Proyecto', 'BU', 'Empresa', 'Amount Total', 'Gross Margin', 'Costo de Venta']]
        
        # Crear DataFrame de totales
        totals_data = []
        for month in month_columns:
            # Sumar valores del DataFrame filtrado
            month_total = df_cost_filtered[month].sum() if month in df_cost_filtered.columns else 0
            if month_total > 0:  # Solo incluir meses con valores
                totals_data.append({
                    'Mes': month,
                    'Total': month_total
                })
        
        if totals_data:
            df_totals = pd.DataFrame(totals_data)
            
            # Crear gráfico de barras horizontal con totales
            fig_totals = px.bar(
                df_totals,
                x='Total',
                y='Mes',
                orientation='h',
                title='Totales Mensuales de Costo de Venta',
                labels={'Total': 'Costo Total ($)', 'Mes': 'Período'},
                text='Total'
            )
            
            fig_totals.update_traces(
                texttemplate='$%{text:,.0f}',
                textposition='outside',
                marker_color='#FCB72F',
                textfont_size=10
            )
            
            fig_totals.update_layout(
                height=max(300, len(df_totals) * 35),
                title_font_size=16,
                title_font_color='#1f4e79',
                font=dict(family="Arial, sans-serif", size=11, color="#2c3e50"),
                plot_bgcolor='rgba(0,0,0,0)',
                paper_bgcolor='rgba(0,0,0,0)',
                showlegend=False,
                xaxis_title="Total Costo de Venta ($)",
                yaxis_title="",
                yaxis={'categoryorder':'total ascending'}
            )
            
            st.plotly_chart(fig_totals, use_container_width=True)
            
            # Mostrar tabla con totales
            st.markdown("##### 📋 Tabla de Totales")
            df_totals['Total'] = df_totals['Total'].apply(lambda x: f"${x:,.2f}")
            st.dataframe(
                df_totals,
                use_container_width=True,
                hide_index=True
            )
            
            # Mostrar gran total
            grand_total = sum([item['Total'] for item in totals_data])
            
            col_total1, col_total2, col_total3 = st.columns(3)
            with col_total1:
                st.metric("💰 Gran Total Costo", f"${grand_total:,.2f}")
            
            with col_total2:
                if 'total_amount' in cost_table:
                    st.metric("📊 Total Amount", f"${cost_table['total_amount']:,.2f}")
            
            with col_total3:
                if 'total_gross_margin' in cost_table:
                    st.metric("📈 Total Gross Margin", f"${cost_table['total_gross_margin']:,.2f}")
    
    def _render_charts(self, summary, billing_events):
        """Renderiza los gráficos del forecast."""
        create_section_header("Visualizaciones", "Análisis gráfico del forecast", "📈")
        
        # Filtros independientes para cada visualización
        st.markdown("#### 🎯 Filtros de Visualización")
        col_filter1, col_filter2, col_filter3 = st.columns(3)
        
        with col_filter1:
            # Obtener lista de empresas disponibles
            available_companies = sorted(list(set([e.company for e in billing_events if e.company])))
            company_filter = st.selectbox(
                "Filtrar por Empresa",
                ['Todas'] + available_companies,
                key="chart_company_filter"
            )
        
        with col_filter2:
            # Obtener lista de BUs disponibles
            available_bus = list(summary.bu_distribution.keys())
            monthly_bu_filter = st.selectbox(
                "Filtrar por BU",
                ['Todas'] + available_bus,
                key="monthly_bu_filter"
            )
        
        #with col_filter2:
        #    temporal_bu_filter = st.selectbox(
        #        "Filtrar Evolución Temporal por BU",
        #        ['Todas'] + available_bus,
        #        key="temporal_bu_filter"
        #    )
        
        col1, col2 = st.columns(2)
        
        with col1:
            # Gráfico de distribución mensual con filtros independientes
            st.markdown("##### 📊 Distribución Mensual")
            
            # Aplicar filtros
            filtered_events = billing_events
            filter_labels = []
            
            if company_filter != 'Todas':
                filtered_events = [e for e in filtered_events if e.company == company_filter]
                filter_labels.append(company_filter)
            
            if monthly_bu_filter != 'Todas':
                filtered_events = [e for e in filtered_events if e.bu.value == monthly_bu_filter]
                filter_labels.append(monthly_bu_filter)
            
            if len(filter_labels) == 0:
                # Mostrar todos los datos
                monthly_data = pd.DataFrame([
                    {'Mes': month, 'Monto': amount}
                    for month, amount in summary.monthly_distribution.items()
                ])
                monthly_title_suffix = ""
            else:
                # Calcular distribución mensual filtrada
                monthly_filtered = {}
                for event in filtered_events:
                    month = event.month_year
                    monthly_filtered[month] = monthly_filtered.get(month, 0) + event.amount_adjusted
                
                monthly_data = pd.DataFrame([
                    {'Mes': month, 'Monto': amount}
                    for month, amount in monthly_filtered.items()
                ])
                monthly_title_suffix = f" - {' / '.join(filter_labels)}"
            
            if not monthly_data.empty:
                fig_monthly = px.bar(
                    monthly_data,
                    x='Mes',
                    y='Monto',
                    title=f'Ingresos por Mes{monthly_title_suffix}',
                    labels={'Monto': 'Monto ($)', 'Mes': 'Período'},
                    color='Monto',
                    color_continuous_scale=[[0, '#E3F2FD'], [0.5, '#2E86AB'], [1, '#1f4e79']]
                )
                fig_monthly.update_layout(
                    xaxis_tickangle=45,
                    showlegend=False,
                    yaxis_title="Ingresos Proyectados ($)",
                    title_font_size=16,
                    title_font_color='#1f4e79',
                    plot_bgcolor='rgba(0,0,0,0)',
                    paper_bgcolor='rgba(0,0,0,0)',
                    font=dict(family="Arial, sans-serif", size=12, color="#2c3e50")
                )
                
                # Agregar valores en las barras
                fig_monthly.update_traces(
                    texttemplate='$%{y:,.2f}',
                    textposition='outside',
                    textfont_size=10,
                    textfont_color='#1f4e79'
                )
                
                st.plotly_chart(fig_monthly, use_container_width=True)
                
                # Mostrar total del filtro
                total_monthly = monthly_data['Monto'].sum()
                filter_text = ' / '.join(filter_labels) if filter_labels else 'Total'
                st.info(f"💰 Total {filter_text}: {fmt_currency(total_monthly, decimals=2)}")
            else:
                filter_text = ' / '.join(filter_labels) if filter_labels else 'este filtro'
                st.warning(f"No hay datos mensuales para mostrar para {filter_text}")
        
        with col2:
            # Gráfico de distribución por BU
            bu_data = pd.DataFrame([
                {'BU': bu, 'Monto': amount}
                for bu, amount in summary.bu_distribution.items()
            ])
            
            if not bu_data.empty:
                # Agregar iconos a los nombres de BU
                bu_data['BU_Display'] = bu_data['BU'].apply(
                    lambda x: f"{format_business_unit_icon(x)} {x}"
                )
                
                fig_bu = px.pie(
                    bu_data,
                    values='Monto',
                    names='BU_Display',
                    title='Distribución por Unidad de Negocio',
                    color_discrete_sequence=['#1f4e79', '#2E86AB', '#40E0D0', '#A23B72', '#28a745']
                )
                fig_bu.update_layout(
                    title_font_size=16,
                    title_font_color='#1f4e79',
                    font=dict(family="Arial, sans-serif", size=12, color="#2c3e50"),
                    plot_bgcolor='rgba(0,0,0,0)',
                    paper_bgcolor='rgba(0,0,0,0)'
                )
                fig_bu.update_traces(
                    textposition='inside',
                    textinfo='percent+label',
                    hovertemplate='<b>%{label}</b><br>Monto: $%{value:,.2f}<br>Porcentaje: %{percent}<extra></extra>'
                )
                st.plotly_chart(fig_bu, use_container_width=True)
        
        # Gráfico de distribución por Empresa
        st.markdown("#### 🏢 Distribución por Empresa")
        
        # Calcular distribución por empresa
        company_distribution = {}
        for event in billing_events:
            company = event.company if event.company else 'Sin Clasificar'
            company_distribution[company] = company_distribution.get(company, 0) + event.amount_adjusted
        
        company_data = pd.DataFrame([
            {'Empresa': company, 'Monto': amount}
            for company, amount in company_distribution.items()
        ])
        
        if not company_data.empty:
            # Crear gráfico de barras para empresas
            fig_company = px.bar(
                company_data,
                x='Empresa',
                y='Monto',
                title='Distribución por Empresa (LLC, SAPI)',
                labels={'Monto': 'Monto ($)', 'Empresa': 'Tipo de Empresa'},
                color='Empresa',
                color_discrete_map={
                    'LLC': '#1f4e79',
                    'SAPI': '#2E86AB', 
                    'Sin Clasificar': '#cccccc'
                }
            )
            fig_company.update_layout(
                title_font_size=16,
                title_font_color='#1f4e79',
                font=dict(family="Arial, sans-serif", size=12, color="#2c3e50"),
                plot_bgcolor='rgba(0,0,0,0)',
                paper_bgcolor='rgba(0,0,0,0)',
                showlegend=False,
                yaxis_title="Ingresos Proyectados ($)"
            )
            fig_company.update_traces(
                texttemplate='$%{y:,.0f}',
                textposition='outside',
                textfont_size=12,
                textfont_color='#1f4e79'
            )
            st.plotly_chart(fig_company, use_container_width=True)
            
            # Mostrar métricas de distribución
            col_comp1, col_comp2, col_comp3 = st.columns(3)
            
            for idx, (company, amount) in enumerate(company_distribution.items()):
                col = [col_comp1, col_comp2, col_comp3][idx % 3]
                with col:
                    percentage = (amount / sum(company_distribution.values())) * 100
                    icon = "🏢" if company == "LLC" else "🏭" if company == "SAPI" else "⚪"
                    st.metric(
                        f"{icon} {company}",
                        f"${amount:,.0f}",
                        f"{percentage:.1f}%"
                    )
        
        # Gráfico de línea temporal
        st.markdown("#### 📅 Evolución Temporal")
        
        # Preparar datos para gráfico de línea (usando filtros aplicados)
        if len(filter_labels) == 0:
            # Usar datos completos
            temporal_monthly_distribution = summary.monthly_distribution
            temporal_title_suffix = ""
        else:
            # Usar datos filtrados (mismos filtros que distribución mensual)
            temporal_monthly_distribution = {}
            for event in filtered_events:
                month = event.month_year
                temporal_monthly_distribution[month] = temporal_monthly_distribution.get(month, 0) + event.amount_adjusted
            temporal_title_suffix = f" - {' / '.join(filter_labels)}"
        
        temporal_monthly_cumulative = []
        temporal_cumulative_amount = 0
        
        for month in sorted(temporal_monthly_distribution.keys()):
            temporal_cumulative_amount += temporal_monthly_distribution[month]
            temporal_monthly_cumulative.append({
                'Mes': month,
                'Mensual': temporal_monthly_distribution[month],
                'Acumulado': temporal_cumulative_amount
            })
        
        if temporal_monthly_cumulative:
            temporal_cumulative_df = pd.DataFrame(temporal_monthly_cumulative)
            
            fig_timeline = go.Figure()
            
            # Línea mensual
            fig_timeline.add_trace(go.Scatter(
                x=temporal_cumulative_df['Mes'],
                y=temporal_cumulative_df['Mensual'],
                mode='lines+markers',
                name='Ingresos Mensuales',
                line=dict(color='#2E86AB', width=4),
                marker=dict(size=10, color='#2E86AB'),
                fill=None
            ))
            
            # Línea acumulada (misma escala)
            fig_timeline.add_trace(go.Scatter(
                x=temporal_cumulative_df['Mes'],
                y=temporal_cumulative_df['Acumulado'],
                mode='lines+markers',
                name='Ingresos Acumulados',
                line=dict(color='#A23B72', width=4, dash='dot'),
                marker=dict(size=10, color='#A23B72')
            ))
            
            fig_timeline.update_layout(
                title=f'Evolución Temporal del Forecast{temporal_title_suffix}',
                xaxis_title='Período',
                yaxis_title='Monto ($)',
                xaxis_tickangle=45,
                hovermode='x unified',
                legend=dict(
                    orientation="h",
                    yanchor="bottom",
                    y=1.02,
                    xanchor="right",
                    x=1
                ),
                plot_bgcolor='rgba(0,0,0,0)',
                paper_bgcolor='rgba(0,0,0,0)',
                title_font_size=16,
                title_font_color='#1f4e79',
                font=dict(family="Arial, sans-serif", size=12, color="#2c3e50"),
                showlegend=True,
                yaxis=dict(
                    tickformat='$,.0f',
                    gridcolor='rgba(128,128,128,0.2)'
                ),
                xaxis=dict(
                    gridcolor='rgba(128,128,128,0.2)'
                )
            )
            
            st.plotly_chart(fig_timeline, use_container_width=True)
            
            # Mostrar total acumulado del filtro
            total_temporal = temporal_cumulative_df['Acumulado'].iloc[-1]
            filter_text = ' / '.join(filter_labels) if filter_labels else 'Total'
            st.info(f"💰 Total Acumulado {filter_text}: ${total_temporal:,.2f}")
        else:
            filter_text = ' / '.join(filter_labels) if filter_labels else 'este filtro'
            st.warning(f"No hay datos temporales para mostrar para {filter_text}")
    
    def _render_details(self, billing_events):
        """Renderiza los detalles de eventos de facturación."""
        st.markdown("### 📋 Detalles de Eventos de Facturación")
        
        if not billing_events:
            st.warning("No hay eventos para mostrar")
            return
        
        # Crear DataFrame con detalles
        details_data = []
        for event in billing_events:
            details_data.append({
                'Proyecto': event.opportunity_name,
                'BU': event.bu.value,
                'Empresa': event.company if event.company else 'Sin Clasificar',
                'Región': event.region if event.region else 'N/A',
                'Etapa': event.stage.value,
                'Fecha': event.date.strftime('%d/%m/%Y'),
                'Mes': event.month_year,
                'Monto Original': event.amount,
                'Monto Ajustado': event.amount_adjusted,
                'Probabilidad': event.probability,  # Mantener como decimal para AG-Grid
                'Lead Time': f"{event.lead_time_adjusted:.1f} sem"
            })
        
        details_df = pd.DataFrame(details_data)
        
        # Filtros avanzados
        st.markdown("#### 🎯 Filtros de Eventos")
        col1, col2, col3, col4, col5 = st.columns(5)
        
        with col1:
            company_filter = st.selectbox(
                "Filtrar por Empresa",
                ['Todas'] + sorted(list(details_df['Empresa'].unique())),
                key="details_company_filter"
            )
        
        with col2:
            bu_filter = st.selectbox(
                "Filtrar por BU",
                ['Todas'] + sorted(list(details_df['BU'].unique())),
                key="details_bu_filter"
            )
        
        with col3:
            stage_filter = st.selectbox(
                "Filtrar por Etapa",
                ['Todas'] + sorted(list(details_df['Etapa'].unique())),
                key="details_stage_filter"
            )
        
        with col4:
            month_filter = st.selectbox(
                "Filtrar por Mes",
                ['Todos'] + sorted(list(details_df['Mes'].unique())),
                key="details_month_filter"
            )
        
        with col5:
            show_selection = st.checkbox("Mostrar solo seleccionados", key="show_selected_details")
        
        # Aplicar filtros
        filtered_df = details_df.copy()
        
        if company_filter != 'Todas':
            filtered_df = filtered_df[filtered_df['Empresa'] == company_filter]
        
        if bu_filter != 'Todas':
            filtered_df = filtered_df[filtered_df['BU'] == bu_filter]
        
        if stage_filter != 'Todas':
            filtered_df = filtered_df[filtered_df['Etapa'] == stage_filter]
        
        if month_filter != 'Todos':
            filtered_df = filtered_df[filtered_df['Mes'] == month_filter]
        
        # Configurar AG-Grid para detalles
        gb_details = AGGridConfigurator.configure_details_table(filtered_df)
        
        # Renderizar AG-Grid de detalles
        grid_config_details = GRID_CONFIGS['details'].copy()
        grid_config_details['height'] = AGGridConfigurator.get_grid_height(len(filtered_df), 500)
        
        st.markdown("#### 📋 Tabla Interactiva de Eventos de Facturación")
        
        details_grid_response = AgGrid(
            filtered_df,
            gridOptions=gb_details.build(),
            height=grid_config_details['height'],
            theme=grid_config_details['theme'],
            allow_unsafe_jscode=grid_config_details['allow_unsafe_jscode'],
            update_mode=grid_config_details['update_mode'],
            fit_columns_on_grid_load=grid_config_details['fit_columns_on_grid_load'],
            enable_enterprise_modules=grid_config_details['enable_enterprise_modules']
        )
        
        # Estadísticas de la tabla de detalles usando GridResponseHandler
        details_handler = GridResponseHandler(details_grid_response)
        
        if details_handler.has_data:
            col_det1, col_det2, col_det3, col_det4 = st.columns(4)
            
            with col_det1:
                st.metric("📊 Eventos", details_handler.total_rows)
            
            with col_det2:
                if 'Monto Ajustado' in details_handler.data_df.columns:
                    total_amount = details_handler.data_df['Monto Ajustado'].sum()
                    st.metric("💰 Total", f"${total_amount:,.0f}")
            
            with col_det3:
                st.metric("✅ Seleccionados", details_handler.selected_count)
            
            with col_det4:
                if 'Proyecto' in details_handler.data_df.columns:
                    unique_projects = details_handler.data_df['Proyecto'].nunique()
                    st.metric("🏗️ Proyectos", unique_projects)
            
            # Controles de exportación para detalles
            st.markdown("#### 🔧 Controles de Exportación")
            col_exp1, col_exp2, col_exp3 = st.columns(3)
            
            with col_exp1:
                if st.button("📥 Exportar Eventos Visibles", key="export_details_visible"):
                    try:
                        export_data = details_handler.export_data("csv")
                        st.download_button(
                            label="⬇️ Descargar CSV Eventos",
                            data=export_data,
                            file_name=f"eventos_facturacion_{datetime.now().strftime('%Y%m%d')}.csv",
                            mime="text/csv"
                        )
                    except Exception as e:
                        st.error(f"Error al exportar eventos: {str(e)}")
            
            with col_exp2:
                if details_handler.has_selection:
                    if st.button("📥 Exportar Eventos Seleccionados", key="export_details_selected"):
                        try:
                            export_data = details_handler.export_selected("csv")
                            st.download_button(
                                label="⬇️ Descargar Selección",
                                data=export_data,
                                file_name=f"eventos_seleccionados_{datetime.now().strftime('%Y%m%d')}.csv",
                                mime="text/csv"
                            )
                        except Exception as e:
                            st.error(f"Error al exportar selección: {str(e)}")
                else:
                    st.info("💡 Selecciona eventos para exportar")
            
            with col_exp3:
                # Resumen de filtros aplicados
                filters_applied = []
                if company_filter != 'Todas':
                    filters_applied.append(f"Empresa: {company_filter}")
                if bu_filter != 'Todas':
                    filters_applied.append(f"BU: {bu_filter}")
                if stage_filter != 'Todas':
                    filters_applied.append(f"Etapa: {stage_filter}")
                if month_filter != 'Todos':
                    filters_applied.append(f"Mes: {month_filter}")
                
                if filters_applied:
                    st.info(f"🔍 Filtros: {', '.join(filters_applied)}")
                else:
                    st.info("📊 Vista completa de eventos")
    
    def _render_analysis(self, billing_events):
        """Renderiza análisis adicionales."""
        st.markdown("### 🎯 Análisis de Riesgo")
        
        # Generar análisis de riesgo
        risk_analysis = self.report_generator.generate_risk_analysis(billing_events)
        
        if 'message' in risk_analysis:
            st.warning(risk_analysis['message'])
            return
        
        # Análisis por probabilidad
        st.markdown("#### 📊 Distribución por Riesgo")
        
        prob_dist = risk_analysis['probability_distribution']
        
        col1, col2, col3 = st.columns(3)
        
        with col1:
            st.metric(
                "🟢 Bajo Riesgo",
                f"${prob_dist['low_risk']['amount']:,.0f}",
                f"{prob_dist['low_risk']['count']} eventos"
            )
        
        with col2:
            st.metric(
                "🟡 Riesgo Medio",
                f"${prob_dist['medium_risk']['amount']:,.0f}",
                f"{prob_dist['medium_risk']['count']} eventos"
            )
        
        with col3:
            st.metric(
                "🔴 Alto Riesgo",
                f"${prob_dist['high_risk']['amount']:,.0f}",
                f"{prob_dist['high_risk']['count']} eventos"
            )
        
        # Análisis de concentración
        st.markdown("#### 🎯 Análisis de Concentración")
        
        conc_risk = risk_analysis['concentration_risk']
        
        if conc_risk['is_concentrated']:
            st.warning(f"⚠️ Alta concentración detectada: {conc_risk['max_bu_concentration']:.1%} en una sola BU")
        else:
            st.success(f"✅ Distribución balanceada: máxima concentración {conc_risk['max_bu_concentration']:.1%}")
        
        # Gráfico de concentración por BU
        conc_data = pd.DataFrame([
            {'BU': bu, 'Monto': amount, 'Porcentaje': amount/sum(conc_risk['bu_distribution'].values())*100}
            for bu, amount in conc_risk['bu_distribution'].items()
        ])
        
        fig_conc = px.bar(
            conc_data,
            x='BU',
            y='Porcentaje',
            title='Concentración por Unidad de Negocio (%)',
            labels={'Porcentaje': 'Porcentaje (%)', 'BU': 'Unidad de Negocio'}
        )
        
        st.plotly_chart(fig_conc, use_container_width=True)
    
    def _render_kpi_billing_table(self):
        """Renderiza la tabla de KPIs PM-008."""
        if not hasattr(st.session_state, 'kpi_results'):
            st.warning("⚠️ No hay datos de KPIs disponibles. Por favor, carga y procesa el archivo de KPIs PM-008.")
            
            with st.expander("📖 ¿Cómo usar esta función?"):
                st.markdown("""
                ### Pasos para cargar KPIs PM-008:
                
                1. **Sube el archivo**: En la barra lateral, busca "Subir archivo de KPI's PM-008"
                2. **Procesa los datos**: Haz clic en el botón "📊 Procesar KPIs PM-008"
                3. **Visualiza**: La tabla aparecerá aquí automáticamente
                
                ### Información que se mostrará:
                - Proyectos con status "Abierto" y "On Hold"
                - Montos de facturación por mes
                - Calculados como: Total de PO × % Facturación
                - Ubicados en el mes de probable facturación
                """)
            return
        
        kpi_results = st.session_state.kpi_results
        
        if not kpi_results['data']:
            st.info("ℹ️ No hay proyectos con status 'Abierto' u 'On Hold' en el archivo de KPIs.")
            return
        
        create_section_header("KPIs PM-008 - Billing", "Proyectos activos con facturación probable", "📋")
        
        # Mostrar resumen
        summary = kpi_results['summary']
        col1, col2, col3, col4 = st.columns(4)
        
        with col1:
            st.metric("📋 Proyectos Activos", summary['total_projects'])
        
        with col2:
            st.metric("💰 Total Billing", f"${summary['total_billing']:,.2f}")
        
        with col3:
            st.metric("📊 Total PO", f"${summary['total_po']:,.2f}")
        
        with col4:
            if summary['status_distribution']:
                abierto = summary['status_distribution'].get('Abierto', 0)
                on_hold = summary['status_distribution'].get('On Hold', 0)
                st.metric("🚦 Abierto / On Hold", f"{abierto} / {on_hold}")
        
        # Convertir a DataFrame
        df_kpis = pd.DataFrame(kpi_results['data'])
        
        # Mostrar warning si hay proyectos con Costo de Venta TBD
        if summary.get('tbd_projects') and len(summary['tbd_projects']) > 0:
            st.warning(f"⚠️ **{len(summary['tbd_projects'])} proyectos con Costo de Venta TBD (pendiente por definir):**")
            
            # Mostrar lista de proyectos
            tbd_list = "\n".join([f"- {project}" for project in summary['tbd_projects']])
            with st.expander(f"📋 Ver {len(summary['tbd_projects'])} proyectos con costo TBD"):
                st.markdown(tbd_list)
                st.info("💡 Estos proyectos se muestran con Costo de Venta = $0.00 hasta que se defina el valor.")
        
        # Filtros
        st.markdown("#### 🔍 Filtros")
        col_filter1, col_filter2, col_filter3, col_filter4 = st.columns(4)
        
        with col_filter1:
            location_options = ['Todas'] + sorted(df_kpis['Location'].dropna().unique().tolist())
            selected_location = st.selectbox("Filtrar por Location:", location_options, key="kpi_location_filter")
        
        with col_filter2:
            bu_options = ['Todas'] + sorted(df_kpis['BU'].dropna().unique().tolist())
            selected_bu = st.selectbox("Filtrar por BU:", bu_options, key="kpi_bu_filter")
        
        with col_filter3:
            status_options = ['Todos'] + sorted(df_kpis['Status'].dropna().unique().tolist())
            selected_status = st.selectbox("Filtrar por Status:", status_options, key="kpi_status_filter")
        
        with col_filter4:
            show_grouping = st.checkbox("Agrupar por BU", value=False, key="kpi_group_by_bu")
        
        # Aplicar filtros
        df_filtered = df_kpis.copy()
        
        if selected_location != 'Todas':
            df_filtered = df_filtered[df_filtered['Location'] == selected_location]
        
        if selected_bu != 'Todas':
            df_filtered = df_filtered[df_filtered['BU'] == selected_bu]
        
        if selected_status != 'Todos':
            df_filtered = df_filtered[df_filtered['Status'] == selected_status]
        
        # Mostrar panel de totales separado (siempre visible)
        self._render_totals_panel(df_filtered, "TOTALES KPIs BILLING", color='#FFF9C4')
        
        # Configurar AG-Grid (SIN fila de totales)
        gb = AGGridConfigurator.configure_forecast_table(df_filtered)
        
        # Configurar agrupación
        if selected_bu != 'Todas' or not show_grouping:
            gb.configure_column("BU", rowGroup=False, hide=False)
            gb.configure_grid_options(groupDefaultExpanded=0)
        else:
            gb.configure_grid_options(groupDefaultExpanded=1)
        
        # Renderizar tabla
        grid_config = GRID_CONFIGS['forecast_main'].copy()
        grid_config['height'] = AGGridConfigurator.get_grid_height(len(df_filtered), 600)
        
        st.markdown("⭐️️ 📋 Tabla de Billing KPIs")
        st.info("💡 Los montos se muestran en el mes de **probable fecha de facturación**. Los totales se muestran en el panel amarillo arriba.")
        
        grid_response = AgGrid(
            df_filtered,
            gridOptions=gb.build(),
            height=grid_config['height'],
            theme=grid_config['theme'],
            allow_unsafe_jscode=grid_config['allow_unsafe_jscode'],
            update_mode=grid_config['update_mode'],
            fit_columns_on_grid_load=grid_config['fit_columns_on_grid_load'],
            enable_enterprise_modules=grid_config['enable_enterprise_modules']
        )
        
        # Estadísticas
        kpi_handler = GridResponseHandler(grid_response)
        
        if kpi_handler.has_data:
            # Ya no hay fila de totales
            df_metrics = kpi_handler.data_df.copy()
            
            st.markdown("#### 📊 Estadísticas")
            col_stat1, col_stat2, col_stat3 = st.columns(3)
            
            with col_stat1:
                st.metric("📋 Proyectos Mostrados", len(df_metrics))
            
            with col_stat2:
                if 'Total PO' in df_metrics.columns:
                    total_po_filtered = pd.to_numeric(df_metrics['Total PO'], errors='coerce').sum()
                    st.metric("💰 Total PO (Filtrado)", f"${total_po_filtered:,.2f}")
            
            with col_stat3:
                # Calcular total de billing de las columnas de meses
                month_cols = [col for col in df_metrics.columns 
                            if col not in ['Proyecto', 'BU', 'Status', 'Customer', 'Total PO', '% Facturación']]
                if month_cols:
                    total_billing_filtered = 0
                    for col in month_cols:
                        total_billing_filtered += pd.to_numeric(df_metrics[col], errors='coerce').sum()
                    st.metric("📈 Total Billing (Filtrado)", f"${total_billing_filtered:,.2f}")
            
            # Exportación
            st.markdown("#### 🔧 Exportar Datos")
            col_export1, col_export2 = st.columns(2)
            
            with col_export1:
                if st.button("📥 Exportar a Excel", key="export_kpi_excel"):
                    try:
                        df_export = df_filtered.copy()
                        buffer = self._export_to_excel_with_format(df_export, 'KPIs PM-008')
                        st.download_button(
                            label="⬇️ Descargar Excel",
                            data=buffer.getvalue(),
                            file_name=f"kpis_pm008_{datetime.now().strftime('%Y%m%d')}.xlsx",
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                        )
                    except Exception as e:
                        st.error(f"Error al exportar: {str(e)}")
            
            with col_export2:
                if st.button("📥 Exportar a CSV", key="export_kpi_csv"):
                    try:
                        df_export = df_filtered.copy()
                        csv_data = df_export.to_csv(index=False)
                        st.download_button(
                            label="⬇️ Descargar CSV",
                            data=csv_data,
                            file_name=f"kpis_pm008_{datetime.now().strftime('%Y%m%d')}.csv",
                            mime="text/csv"
                        )
                    except Exception as e:
                        st.error(f"Error al exportar: {str(e)}")
    
    def _render_kpi_cost_of_sale_table(self):
        """Renderiza la tabla de Costo de Venta de KPIs PM-008."""
        if not hasattr(st.session_state, 'kpi_results'):
            st.warning("⚠️ No hay datos de KPIs disponibles. Por favor, carga y procesa el archivo de KPIs PM-008 primero.")
            
            with st.expander("📖 ¿Cómo usar esta función?"):
                st.markdown("""
                ### Pasos para cargar KPIs PM-008:
                
                1. **Sube el archivo**: En la barra lateral, busca "Subir archivo de KPI's PM-008"
                2. **Procesa los datos**: Haz clic en el botón "📊 Procesar KPIs PM-008"
                3. **Visualiza**: La tabla de costo de venta aparecerá aquí automáticamente
                
                ### Información que se mostrará:
                - Costo de venta por proyecto
                - **Ubicado solo en el último mes de facturación del proyecto**
                - Filtros por Location, BU y Status
                
                ### Nota importante:
                - Si un proyecto tiene múltiples eventos de facturación, el costo de venta 
                  aparecerá **solo una vez** en el último mes, no en cada mes.
                """)
            return
        
        kpi_results = st.session_state.kpi_results
        
        if not kpi_results['data']:
            st.info("ℹ️ No hay proyectos con costo de venta en el archivo de KPIs.")
            return
        
        create_section_header("KPIs PM-008 - Costo de Venta", "Costo de venta de proyectos activos", "💵")
        
        # Convertir a DataFrame
        df_kpis = pd.DataFrame(kpi_results['data'])
        
        # Filtros
        st.markdown("#### 🔍 Filtros")
        col_filter1, col_filter2, col_filter3, col_filter4 = st.columns(4)
        
        with col_filter1:
            location_options = ['Todas'] + sorted(df_kpis['Location'].dropna().unique().tolist())
            selected_location_cost = st.selectbox("Filtrar por Location:", location_options, key="kpi_cost_location_filter")
        
        with col_filter2:
            bu_options = ['Todas'] + sorted(df_kpis['BU'].dropna().unique().tolist())
            selected_bu_cost = st.selectbox("Filtrar por BU:", bu_options, key="kpi_cost_bu_filter")
        
        with col_filter3:
            status_options = ['Todos'] + sorted(df_kpis['Status'].dropna().unique().tolist())
            selected_status_cost = st.selectbox("Filtrar por Status:", status_options, key="kpi_cost_status_filter")
        
        with col_filter4:
            show_grouping_cost = st.checkbox("Agrupar por BU", value=False, key="kpi_cost_group_by_bu")
        
        # Aplicar filtros
        df_filtered = df_kpis.copy()
        
        if selected_location_cost != 'Todas':
            df_filtered = df_filtered[df_filtered['Location'] == selected_location_cost]
        
        if selected_bu_cost != 'Todas':
            df_filtered = df_filtered[df_filtered['BU'] == selected_bu_cost]
        
        if selected_status_cost != 'Todos':
            df_filtered = df_filtered[df_filtered['Status'] == selected_status_cost]
        
        # Crear tabla solo con costo de venta
        # Identificar columnas de meses
        month_cols = [col for col in df_filtered.columns 
                     if col not in ['Proyecto', 'BU', 'Location', 'Status', 'Customer', 'Total PO', '% Facturación', 'Costo de Venta']]
        
        # Crear nueva tabla con costo de venta distribuido por mes
        df_cost = df_filtered[['Proyecto', 'BU', 'Location', 'Status', 'Customer', 'Costo de Venta']].copy()
        
        # Agregar columnas de meses con costo de venta
        # El costo de venta se muestra SOLO en el ÚLTIMO mes donde hay facturación
        for month in month_cols:
            df_cost[month] = 0
        
        # Para cada proyecto, encontrar el último mes con facturación y asignar costo total ahí
        for idx in df_filtered.index:
            # Encontrar el último mes donde hay facturación (valor > 0)
            last_billing_month = None
            for month in month_cols:
                if df_filtered.loc[idx, month] > 0:
                    last_billing_month = month
            
            # Asignar el costo de venta TOTAL solo al último mes
            if last_billing_month:
                df_cost.loc[idx, last_billing_month] = df_filtered.loc[idx, 'Costo de Venta']
        
        # Mostrar panel de totales separado (siempre visible)
        self._render_totals_panel(df_cost, "TOTALES COSTO VENTA KPIs", color='#E8F5E9')
        
        # Configurar AG-Grid (SIN fila de totales)
        gb = AGGridConfigurator.configure_forecast_table(df_cost)
        
        # Configurar agrupación
        if selected_bu_cost != 'Todas' or not show_grouping_cost:
            gb.configure_column("BU", rowGroup=False, hide=False)
            gb.configure_grid_options(groupDefaultExpanded=0)
        else:
            gb.configure_grid_options(groupDefaultExpanded=1)
        
        # Estilo para celdas con costo de venta
        cell_style_jscode = JsCode("""
        function(params) {
            if (params.value && params.value > 0) {
                return {
                    'backgroundColor': '#FCB72F',
                    'color': '#000000',
                    'fontWeight': 'bold'
                }
            }
            return null;
        }
        """)
        
        # Aplicar estilo a columnas de meses
        for month in month_cols:
            gb.configure_column(month, cellStyle=cell_style_jscode)
        
        # Renderizar tabla
        grid_config = GRID_CONFIGS['forecast_main'].copy()
        grid_config['height'] = AGGridConfigurator.get_grid_height(len(df_cost), 600)
        
        st.markdown("⭐️️ 💵 Tabla de Costo de Venta KPIs")
        st.info("💡 El costo de venta se muestra **SOLO en el último mes de facturación** del proyecto. Las celdas con costo están resaltadas en naranja. Los totales se muestran en el panel verde arriba.")
        
        grid_response = AgGrid(
            df_cost,
            gridOptions=gb.build(),
            height=grid_config['height'],
            theme=grid_config['theme'],
            allow_unsafe_jscode=grid_config['allow_unsafe_jscode'],
            update_mode=grid_config['update_mode'],
            fit_columns_on_grid_load=grid_config['fit_columns_on_grid_load'],
            enable_enterprise_modules=grid_config['enable_enterprise_modules']
        )
        
        # Estadísticas
        cost_handler = GridResponseHandler(grid_response)
        
        if cost_handler.has_data:
            # Ya no hay fila de totales
            df_metrics = cost_handler.data_df.copy()
            
            st.markdown("#### 📊 Estadísticas")
            col_stat1, col_stat2, col_stat3 = st.columns(3)
            
            with col_stat1:
                st.metric("📋 Proyectos Mostrados", len(df_metrics))
            
            with col_stat2:
                if 'Costo de Venta' in df_metrics.columns:
                    total_cost = pd.to_numeric(df_metrics['Costo de Venta'], errors='coerce').sum()
                    st.metric("💰 Total Costo de Venta", f"${total_cost:,.2f}")
            
            with col_stat3:
                # Calcular total de costo de las columnas de meses
                if month_cols:
                    total_cost_monthly = 0
                    for col in month_cols:
                        total_cost_monthly += pd.to_numeric(df_metrics[col], errors='coerce').sum()
                    st.metric("📈 Total Costo Mensual", f"${total_cost_monthly:,.2f}")
            
            # Exportación
            st.markdown("#### 🔧 Exportar Datos")
            col_export1, col_export2 = st.columns(2)
            
            with col_export1:
                if st.button("📥 Exportar a Excel", key="export_kpi_cost_excel"):
                    try:
                        df_export = df_cost.copy()
                        buffer = self._export_to_excel_with_format(df_export, 'Costo Venta KPIs')
                        st.download_button(
                            label="⬇️ Descargar Excel",
                            data=buffer.getvalue(),
                            file_name=f"costo_venta_kpis_pm008_{datetime.now().strftime('%Y%m%d')}.xlsx",
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                        )
                    except Exception as e:
                        st.error(f"Error al exportar: {str(e)}")
            
            with col_export2:
                if st.button("📥 Exportar a CSV", key="export_kpi_cost_csv"):
                    try:
                        df_export = df_cost.copy()
                        csv_data = df_export.to_csv(index=False)
                        st.download_button(
                            label="⬇️ Descargar CSV",
                            data=csv_data,
                            file_name=f"costo_venta_kpis_pm008_{datetime.now().strftime('%Y%m%d')}.csv",
                            mime="text/csv"
                        )
                    except Exception as e:
                        st.error(f"Error al exportar: {str(e)}")
    
    def _render_chatbot(self, results: Dict[str, Any]):
        """Renderiza el chatbot asistente de forecast."""
        from src.chatbot import ForecastChatbot
        
        create_section_header("Asistente de Forecast", "Analiza tus datos con IA", "🤖")
        
        # Inicializar chatbot en session_state si no existe
        if 'chatbot' not in st.session_state:
            st.session_state.chatbot = ForecastChatbot()
        
        if 'chat_messages' not in st.session_state:
            st.session_state.chat_messages = []
        
        chatbot = st.session_state.chatbot
        
        # Sidebar con configuración
        with st.sidebar:
            st.markdown("### ⚙️ Configuración del Chatbot")
            
            # Input para API Key
            api_key_input = st.text_input(
                "OpenAI API Key",
                type="password",
                value=st.session_state.get('openai_api_key', ''),
                help="Ingresa tu API key de OpenAI para habilitar el chatbot"
            )
            
            if api_key_input:
                st.session_state.openai_api_key = api_key_input
                chatbot.set_api_key(api_key_input)
                st.success("✅ API Key configurada")
            
            # Selector de modelo
            model = st.selectbox(
                "Modelo",
                ["gpt-4o-mini", "gpt-4o", "gpt-3.5-turbo"],
                help="gpt-4o-mini es más económico y rápido"
            )
            chatbot.model = model
            
            # Botón para limpiar historial
            if st.button("🗑️ Limpiar Conversación"):
                chatbot.clear_history()
                st.session_state.chat_messages = []
                st.rerun()
            
            st.markdown("---")
            st.markdown("### 💡 Ejemplos de Preguntas")
            st.markdown("""
            - ¿Cuál es el total del forecast?
            - Muéstrame los top 10 proyectos
            - Analiza la BU de FCT
            - ¿Cómo se distribuye por empresa?
            - Busca proyectos de "Microsoft"
            - ¿Cuál es el forecast para los próximos 3 meses?
            - Analiza el costo de venta
            """)
        
        # Verificar si está configurado
        if not chatbot.is_configured():
            st.warning("⚠️ Por favor configura tu API key de OpenAI en la barra lateral para usar el chatbot.")
            st.info("💡 El chatbot usa la API de OpenAI para analizar tus datos de forecast. Necesitas una API key para continuar.")
            
            with st.expander("📖 ¿Cómo obtener una API Key?"):
                st.markdown("""
                1. Ve a [platform.openai.com](https://platform.openai.com)
                2. Crea una cuenta o inicia sesión
                3. Ve a "API Keys" en tu perfil
                4. Crea una nueva API key
                5. Copia la key y pégala en el campo de la barra lateral
                
                **Nota**: El uso de la API tiene costos asociados. Revisa la [página de precios de OpenAI](https://openai.com/pricing).
                """)
            return
        
        # Interfaz de chat
        st.markdown("#### 💬 Conversación")
        
        # Mostrar mensajes del historial
        for message in st.session_state.chat_messages:
            with st.chat_message(message["role"]):
                st.markdown(message["content"])
        
        # Input del usuario
        if prompt := st.chat_input("Escribe tu pregunta aquí..."):
            # Agregar mensaje del usuario
            st.session_state.chat_messages.append({"role": "user", "content": prompt})
            
            # Mostrar mensaje del usuario
            with st.chat_message("user"):
                st.markdown(prompt)
            
            # Generar respuesta
            with st.chat_message("assistant"):
                with st.spinner("Pensando..."):
                    try:
                        response = chatbot.chat(prompt, results)
                        st.markdown(response)
                        
                        # Agregar respuesta al historial
                        st.session_state.chat_messages.append({"role": "assistant", "content": response})
                    except Exception as e:
                        error_msg = f"❌ Error: {str(e)}"
                        st.error(error_msg)
                        st.session_state.chat_messages.append({"role": "assistant", "content": error_msg})
        
        # Información adicional
        with st.expander("ℹ️ Acerca del Chatbot"):
            st.markdown("""
            El **Asistente de Forecast** es un chatbot potenciado por IA que puede:
            
            - 📊 Consultar y analizar tus datos de forecast
            - 🎯 Responder preguntas sobre proyectos, BUs y distribuciones
            - 💰 Analizar costos de venta y márgenes
            - 🔍 Buscar proyectos específicos
            - 📈 Generar análisis y resúmenes
            
            **Funciones Disponibles:**
            - Resumen ejecutivo del forecast
            - Análisis mensual
            - Análisis por Business Unit
            - Top proyectos
            - Distribución por empresa (LLC/SAPI)
            - Análisis de costo de venta
            - Búsqueda de proyectos
            
            El chatbot tiene acceso completo a los datos cargados en esta sesión.
            """)
    
    def _render_processing_info(self, processing_summary, validation_result):
        """Renderiza información del procesamiento."""
        st.markdown("### ⚙️ Información del Procesamiento")
        
        col1, col2 = st.columns(2)
        
        with col1:
            st.markdown("#### 📊 Resumen de Datos")
            st.metric("Registros Originales", processing_summary['original_records'])
            st.metric("Registros Válidos", processing_summary['valid_records'])
            excluded_100 = processing_summary.get('excluded_100_percent', 0)
            st.metric("Excluidos (100%)", excluded_100)
            st.metric("Tasa de Éxito", f"{processing_summary['success_rate']:.1%}")
            st.metric("Ajustes de Lead Time", processing_summary['lead_time_adjustments'])
        
        with col2:
            st.markdown("#### 🎯 Distribución de Datos")
            
            # Distribución por Empresa
            if 'company_distribution' in processing_summary and processing_summary['company_distribution']:
                st.markdown("**Por Empresa:**")
                for company, count in processing_summary['company_distribution'].items():
                    icon = "🏢" if company == "LLC" else "🏭" if company == "SAPI" else "⚪"
                    st.text(f"{icon} {company}: {count} oportunidades")
            
            # Distribución por BU
            st.markdown("**Por BU:**")
            for bu, count in processing_summary['bu_distribution'].items():
                st.text(f"• {bu}: {count} oportunidades")
            
            # Distribución por probabilidad
            st.markdown("**Por Probabilidad:**")
            for prob, count in processing_summary['probability_distribution'].items():
                st.text(f"• {prob:.0%}: {count} oportunidades")
        
        # Mostrar información de parsing del archivo
        if 'parsing_report' in st.session_state.forecast_results:
            parsing_info = st.session_state.forecast_results['parsing_report']
            
            st.markdown("#### 📄 Información de Parsing del Archivo")
            
            col1, col2, col3 = st.columns(3)
            
            with col1:
                st.metric("Fila de Headers Detectada", parsing_info.get('detected_header_row', 'N/A'))
            
            with col2:
                st.metric("Columnas Originales", parsing_info.get('original_columns_count', 0))
            
            with col3:
                parsing_success = "✅ Exitoso" if parsing_info.get('parsing_success', False) else "❌ Con errores"
                st.metric("Estado del Parsing", parsing_success)
            
            # Mostrar mapeos de columnas aplicados
            if parsing_info.get('applied_mappings'):
                st.markdown("**Mapeos de Columnas Aplicados:**")
                for original, normalized in parsing_info['applied_mappings'].items():
                    st.text(f"• '{original}' → '{normalized}'")
            
            # Mostrar si se normalizó PIA
            if parsing_info.get('pia_normalization_applied'):
                st.info("💡 Se detectaron y normalizaron valores de 'Paid in Advance' desde porcentajes a montos")
        
        # Mostrar información sobre datos completados
        st.markdown("#### 🔧 Datos Completados Automáticamente")
        
        # Información sobre Lead Times completados
        if 'lead_time_source' in processing_summary:
            lead_time_stats = processing_summary.get('lead_time_completion', {})
            if lead_time_stats:
                col1, col2, col3 = st.columns(3)
                with col1:
                    st.metric("Lead Times Históricos", lead_time_stats.get('historical', 0))
                with col2:
                    st.metric("Lead Times Estimados", lead_time_stats.get('estimated', 0))
                with col3:
                    st.metric("Lead Times Originales", lead_time_stats.get('original', 0))
        
        # Información sobre Payment Terms completados
        if 'payment_terms_source' in processing_summary:
            payment_terms_stats = processing_summary.get('payment_terms_completion', {})
            if payment_terms_stats:
                col1, col2, col3 = st.columns(3)
                with col1:
                    st.metric("Payment Terms Históricos", payment_terms_stats.get('historical', 0))
                with col2:
                    st.metric("Payment Terms por Defecto", payment_terms_stats.get('default', 0))
                with col3:
                    st.metric("Payment Terms Originales", payment_terms_stats.get('original', 0))
        
        # Mostrar advertencias de validación
        if validation_result.warnings:
            st.markdown("#### ⚠️ Advertencias de Validación")
            for warning in validation_result.warnings[:10]:  # Mostrar máximo 10
                st.warning(warning)
    
    def _export_excel(self):
        """Exporta resultados a Excel."""
        try:
            results = st.session_state.forecast_results
            
            buffer = self.exporter.export_to_excel(
                results['billing_events'],
                results['forecast_table'],
                results['summary']
            )
            
            filename = self.exporter.create_downloadable_filename("forecast")
            
            st.download_button(
                label="📊 Descargar Excel",
                data=buffer.getvalue(),
                file_name=filename,
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
            
            st.success("✅ Archivo Excel preparado para descarga")
            
        except Exception as e:
            st.error(f"❌ Error al exportar Excel: {str(e)}")
    
    def _export_csv(self):
        """Exporta tabla de forecast a CSV."""
        try:
            results = st.session_state.forecast_results
            
            csv_content = self.exporter.export_forecast_table_to_csv(results['forecast_table'])
            
            filename = self.exporter.create_downloadable_filename("forecast").replace('.xlsx', '.csv')
            
            st.download_button(
                label="📄 Descargar CSV",
                data=csv_content,
                file_name=filename,
                mime="text/csv"
            )
            
            st.success("✅ Archivo CSV preparado para descarga")
            
        except Exception as e:
            st.error(f"❌ Error al exportar CSV: {str(e)}")
    
    def _render_forecast_table_low_prob(self, forecast_table, summary):
        """Renderiza la tabla de forecast para oportunidades con probabilidad < 60%."""
        
        # Métricas principales para oportunidades < 60%
        if summary:
            col1, col2, col3, col4 = st.columns(4)
            
            with col1:
                st.metric(
                    "💰 Total Forecast <60%",
                    fmt_currency(summary.total_amount, decimals=2),
                    help="Monto total proyectado con factores de castigo aplicados"
                )
            
            with col2:
                st.metric(
                    "🎯 Oportunidades <60%",
                    f"{summary.total_opportunities:,}",
                    help="Número de oportunidades con probabilidad menor a 60%"
                )
            
            with col3:
                st.metric(
                    "📅 Eventos",
                    f"{summary.total_events:,}",
                    help="Número total de eventos de facturación"
                )
            
            with col4:
                st.metric(
                    "⏱️ Duración",
                    f"{summary.duration_months} meses",
                    help="Duración del forecast en meses"
                )
            
            # Botón de descarga de totales consolidados para <60%
            st.markdown("---")
            col_download, col_info = st.columns([1, 3])
            
            with col_download:
                if st.button("📊 Descargar Reporte Consolidado de Totales <60%", key="download_consolidated_low_prob", help="Descarga Excel con totales por Empresa y BU para oportunidades <60%", use_container_width=True):
                    try:
                        excel_buffer = self._generate_consolidated_totals_excel_low_prob()
                        st.download_button(
                            label="⬇️ Descargar Excel Consolidado <60%",
                            data=excel_buffer.getvalue(),
                            file_name=f"reporte_consolidado_totales_low_prob_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx",
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                            key="download_consolidated_low_prob_btn"
                        )
                    except Exception as e:
                        st.error(f"Error al generar reporte: {str(e)}")
            
            with col_info:
                st.info("📋 Incluye: Hoja 'Totales <60%' con resumen general + Una hoja por cada BU con sus totales específicos")
        
        st.markdown("---")
        
        if len(forecast_table['data']) == 0:
            st.warning("No hay datos de oportunidades con probabilidad < 60%")
            return
        
        df = pd.DataFrame(forecast_table['data'])
        
        # Filtros
        filter_configs = [
            {'column': 'Empresa', 'label': '🏢 Empresa', 'key': 'forecast_low_prob_empresa'},
            {'column': 'BU', 'label': '📋 BU', 'key': 'forecast_low_prob_bu'}
        ]
        
        col_filters, col_controls = st.columns([3, 1])
        
        with col_filters:
            selected_filters = self._render_filters_row(df, filter_configs)
        
        with col_controls:
            show_grouping = st.checkbox("Agrupar por BU", value=True, key="forecast_low_prob_group")
        
        # Aplicar filtros
        df_filtered = df.copy()
        
        for column, value in selected_filters.items():
            if value != 'Todas':
                df_filtered = df_filtered[df_filtered[column] == value]
        
        # Mostrar panel de totales
        self._render_totals_panel(df_filtered, "TOTALES FORECAST <60%")
        
        # Configurar AG-Grid
        gb = AGGridConfigurator.configure_forecast_table(df_filtered)
        
        if selected_filters['BU'] != 'Todas' or not show_grouping:
            gb.configure_column("BU", rowGroup=False, hide=False)
            gb.configure_grid_options(groupDefaultExpanded=0)
        
        # Renderizar AG-Grid
        grid_config = GRID_CONFIGS['forecast_main'].copy()
        grid_config['height'] = AGGridConfigurator.get_grid_height(len(df_filtered), 600)
        
        st.markdown("#### 📉 Tabla de Forecast <60%")
        
        grid_response = AgGrid(
            df_filtered,
            gridOptions=gb.build(),
            height=grid_config['height'],
            theme=grid_config['theme'],
            allow_unsafe_jscode=grid_config['allow_unsafe_jscode'],
            update_mode=grid_config['update_mode'],
            fit_columns_on_grid_load=grid_config['fit_columns_on_grid_load'],
            enable_enterprise_modules=grid_config['enable_enterprise_modules']
        )
        
        # Estadísticas
        grid_handler = GridResponseHandler(grid_response)
        
        if grid_handler.has_data:
            df_forecast_metrics = grid_handler.data_df.copy()
            
            col_stats1, col_stats2, col_stats3 = st.columns(3)
            
            with col_stats1:
                st.metric("📋 Proyectos", len(df_forecast_metrics))
            
            with col_stats2:
                numeric_columns = df_forecast_metrics.select_dtypes(include=['number']).columns
                numeric_columns = [col for col in numeric_columns if col not in ['BU']]
                if len(numeric_columns) > 0:
                    total_forecast = 0
                    for col in numeric_columns:
                        total_forecast += pd.to_numeric(df_forecast_metrics[col], errors='coerce').sum()
                    st.metric("💰 Total Forecast", f"${total_forecast:,.0f}")
            
            with col_stats3:
                unique_bus = df_forecast_metrics['BU'].nunique()
                st.metric("🏢 BUs Activas", unique_bus)
            
            # Exportación
            st.markdown("#### 📥 Exportar Datos")
            self._render_export_buttons(df_filtered, 'forecast_low_prob', 'forecast_low_prob_export')
    
    def _render_cost_of_sale_table_low_prob(self, cost_table, summary):
        """Renderiza la tabla de costo de venta para oportunidades con probabilidad < 60%."""
        
        if len(cost_table['data']) == 0:
            st.warning("No hay datos de costo de venta para oportunidades < 60%")
            return
        
        df = pd.DataFrame(cost_table['data'])
        
        # Métricas principales
        col1, col2, col3 = st.columns(3)
        
        # Excluir fila de totales para métricas
        df_for_metrics = df[df['Proyecto'] != 'TOTAL COSTO'].copy() if 'TOTAL COSTO' in df['Proyecto'].values else df.copy()
        
        with col1:
            st.metric("📋 Proyectos <60%", len(df_for_metrics))
        
        with col2:
            if 'Costo de Venta' in df_for_metrics.columns:
                total_cost = pd.to_numeric(df_for_metrics['Costo de Venta'], errors='coerce').sum()
                st.metric("💸 Total Costo de Venta <60%", f"${total_cost:,.0f}")
        
        with col3:
            if 'Gross Margin' in df_for_metrics.columns:
                total_margin = pd.to_numeric(df_for_metrics['Gross Margin'], errors='coerce').sum()
                st.metric("💰 Total Gross Margin <60%", f"${total_margin:,.0f}")
        
        st.markdown("---")
        
        # Filtros
        filter_configs = [
            {'column': 'Empresa', 'label': '🏢 Empresa', 'key': 'cost_low_prob_empresa'},
            {'column': 'BU', 'label': '📋 BU', 'key': 'cost_low_prob_bu'}
        ]
        
        col_filters, col_controls = st.columns([3, 1])
        
        with col_filters:
            selected_filters = self._render_filters_row(df, filter_configs)
        
        with col_controls:
            show_grouping = st.checkbox("Agrupar por BU", value=True, key="cost_low_prob_group")
        
        # Aplicar filtros
        df_filtered = df.copy()
        
        for column, value in selected_filters.items():
            if value != 'Todas':
                df_filtered = df_filtered[df_filtered[column] == value]
        
        # Mostrar panel de totales
        self._render_totals_panel(df_filtered, "TOTALES COSTO DE VENTA <60%")
        
        # Configurar AG-Grid
        
        gb = AGGridConfigurator.configure_forecast_table(df_filtered)
        
        if selected_filters['BU'] != 'Todas' or not show_grouping:
            gb.configure_column("BU", rowGroup=False, hide=False)
            gb.configure_grid_options(groupDefaultExpanded=0)
        
        # Renderizar AG-Grid
        grid_config = GRID_CONFIGS['forecast_main'].copy()
        grid_config['height'] = AGGridConfigurator.get_grid_height(len(df_filtered), 600)
        
        st.markdown("#### 💸 Tabla de Costo de Venta <60%")
        
        grid_response = AgGrid(
            df_filtered,
            gridOptions=gb.build(),
            height=grid_config['height'],
            theme=grid_config['theme'],
            allow_unsafe_jscode=grid_config['allow_unsafe_jscode'],
            update_mode=grid_config['update_mode'],
            fit_columns_on_grid_load=grid_config['fit_columns_on_grid_load'],
            enable_enterprise_modules=grid_config['enable_enterprise_modules']
        )
        
        # Exportación
        st.markdown("#### 📥 Exportar Datos")
        self._render_export_buttons(df_filtered, 'cost_low_prob', 'cost_low_prob_export')
    
    def _generate_consolidated_totals_excel_low_prob(self):
        """Genera un Excel con totales consolidados por Empresa y BU para oportunidades <60%."""
        from io import BytesIO
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        
        results = st.session_state.forecast_results
        forecast_table = results['forecast_table_low_prob']
        cost_table = results['cost_of_sale_table_low_prob']
        
        df_forecast = pd.DataFrame(forecast_table['data'])
        df_cost = pd.DataFrame(cost_table['data'])
        
        # Excluir filas de totales
        df_forecast = df_forecast[df_forecast['Proyecto'] != 'TOTAL FORECAST'].copy()
        df_cost = df_cost[df_cost['Proyecto'] != 'TOTAL COSTO'].copy()
        
        buffer = BytesIO()
        
        with pd.ExcelWriter(buffer, engine='openpyxl') as writer:
            # Hoja 1: Totales Generales
            totals_data_forecast = []
            totals_data_cost = []
            
            # Por Empresa
            for empresa in df_forecast['Empresa'].unique():
                df_empresa_f = df_forecast[df_forecast['Empresa'] == empresa]
                df_empresa_c = df_cost[df_cost['Empresa'] == empresa]
                
                # Sumar columnas numéricas
                numeric_cols_f = df_empresa_f.select_dtypes(include=['number']).columns
                numeric_cols_f = [col for col in numeric_cols_f if col not in ['BU']]
                total_f = sum([pd.to_numeric(df_empresa_f[col], errors='coerce').sum() for col in numeric_cols_f])
                
                numeric_cols_c = df_empresa_c.select_dtypes(include=['number']).columns
                numeric_cols_c = [col for col in numeric_cols_c if col not in ['BU']]
                total_c = sum([pd.to_numeric(df_empresa_c[col], errors='coerce').sum() for col in numeric_cols_c])
                
                totals_data_forecast.append({'Empresa': empresa, 'Total Forecast <60%': total_f})
                totals_data_cost.append({'Empresa': empresa, 'Total Costo <60%': total_c})
            
            df_totals_f = pd.DataFrame(totals_data_forecast)
            df_totals_c = pd.DataFrame(totals_data_cost)
            df_totals = pd.merge(df_totals_f, df_totals_c, on='Empresa')
            
            df_totals.to_excel(writer, sheet_name='Totales <60%', index=False)
            
            # Formatear hoja de totales
            ws_totals = writer.sheets['Totales <60%']
            for row in ws_totals.iter_rows(min_row=2, max_row=len(df_totals)+1, min_col=2, max_col=3):
                for cell in row:
                    cell.number_format = '$#,##0.00'
            
            # Hojas por BU
            for bu in df_forecast['BU'].unique():
                df_bu_f = df_forecast[df_forecast['BU'] == bu]
                df_bu_c = df_cost[df_cost['BU'] == bu]
                
                # Totales por empresa dentro de cada BU
                bu_totals_f = []
                bu_totals_c = []
                
                for empresa in df_bu_f['Empresa'].unique():
                    df_empresa_f = df_bu_f[df_bu_f['Empresa'] == empresa]
                    df_empresa_c = df_bu_c[df_bu_c['Empresa'] == empresa]
                    
                    numeric_cols_f = df_empresa_f.select_dtypes(include=['number']).columns
                    numeric_cols_f = [col for col in numeric_cols_f if col not in ['BU']]
                    total_f = sum([pd.to_numeric(df_empresa_f[col], errors='coerce').sum() for col in numeric_cols_f])
                    
                    numeric_cols_c = df_empresa_c.select_dtypes(include=['number']).columns
                    numeric_cols_c = [col for col in numeric_cols_c if col not in ['BU']]
                    total_c = sum([pd.to_numeric(df_empresa_c[col], errors='coerce').sum() for col in numeric_cols_c])
                    
                    bu_totals_f.append({'Empresa': empresa, 'Total Forecast': total_f})
                    bu_totals_c.append({'Empresa': empresa, 'Total Costo': total_c})
                
                df_bu_totals_f = pd.DataFrame(bu_totals_f)
                df_bu_totals_c = pd.DataFrame(bu_totals_c)
                df_bu_totals = pd.merge(df_bu_totals_f, df_bu_totals_c, on='Empresa')
                
                sheet_name = f'{bu} <60%'
                df_bu_totals.to_excel(writer, sheet_name=sheet_name, index=False)
                
                # Formatear
                ws_bu = writer.sheets[sheet_name]
                for row in ws_bu.iter_rows(min_row=2, max_row=len(df_bu_totals)+1, min_col=2, max_col=3):
                    for cell in row:
                        cell.number_format = '$#,##0.00'
        
        buffer.seek(0)
        return buffer


# Punto de entrada de la aplicación
if __name__ == "__main__":
    app = ForecastApp()
    app.run()
